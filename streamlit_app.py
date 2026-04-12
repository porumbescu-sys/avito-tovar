from __future__ import annotations

import html
import io
import json
import math
import re
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

st.set_page_config(page_title="Мой Товар", page_icon="📦", layout="wide")

APP_TITLE = "Мой Товар"
DEFAULT_DISCOUNT_1 = 12.0
DEFAULT_DISCOUNT_2 = 20.0
DEFAULT_TEMPLATE1_FOOTER = (
    "Цeна с НДC : +17%\n\n"
    "Работaeм по будням, c 10 дo 18:00. Самовывоз по адресу: Москва, ул. Сущёвский Вал, 5с20\n\n"
    "Еcли пoтрeбуeтся пepeсылкa - oтпpaвляeм толькo Авитo-Яндeкc, Авито-СДЭК или Авито-Авито. Отправляем без наценки."
)

COLUMN_ALIASES = {
    "article": ["Артикул", "артикул", "код", "sku", "артикл", "article"],
    "name": ["Номенклатура", "Наименование", "название", "товар", "name"],
    "brand": [
        "Номенклатура.Производитель",
        "Производитель",
        "бренд",
        "марка",
        "brand",
    ],
    "free_qty": ["Свободно", "Свободный остаток", "остаток", "наличие", "free"],
    "total_qty": ["Всего", "Количество", "всего на складе", "total"],
    "price": ["Цена", "Цена продажи", "Продажа", "розница", "price"],
}

AVITO_COLUMN_ALIASES = {
    "ad_id": ["Номер объявления", "ID объявления", "Номер"],
    "title": ["Название объявления", "Заголовок", "Название"],
    "price": ["Цена"],
    "url": ["Ссылка", "URL", "Ссылка на объявление", "Link"],
}

ARTICLE_REFERENCE_COLUMN_ALIASES = {
    "brand": ["Производитель", "Бренд", "Марка", "brand"],
    "article": ["Артикул", "Короткий артикул", "Наш артикул", "article"],
    "manufacturer_article": ["Артикул производителя", "OEM", "OEM-код", "Код производителя", "manufacturer_article"],
    "name": ["Номенклатура", "Наименование", "Название", "name"],
}

COLOR_KEYWORDS = [
    ("желтый", "желтый"),
    ("yellow", "желтый"),
    ("cyan", "голубой"),
    ("голубой", "голубой"),
    ("синий", "синий"),
    ("blue", "синий"),
    ("magenta", "пурпурный"),
    ("пурпур", "пурпурный"),
    ("фиолет", "пурпурный"),
    ("purple", "пурпурный"),
    ("red", "красный"),
    ("красный", "красный"),
    ("black", "черный"),
    ("черный", "черный"),
    ("чёрный", "черный"),
    ("grey", "серый"),
    ("gray", "серый"),
    ("серый", "серый"),
    ("green", "зеленый"),
    ("зел", "зеленый"),
]

CYRILLIC_ARTICLE_TRANSLATION = str.maketrans({
    "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "У": "Y", "Х": "X",
    "а": "A", "в": "B", "е": "E", "к": "K", "м": "M", "н": "H", "о": "O", "р": "P", "с": "C", "т": "T", "у": "Y", "х": "X",
    "Ё": "E", "ё": "E",
})

ARTICLE_PIECE_RE = re.compile(r"^[A-Za-zА-Яа-я0-9._-]{3,}$")
SERIES_SUFFIX_ORDER = {
    "A": 0,
    "AC": 1,
    "X": 2,
    "XH": 3,
    "XC": 4,
    "Y": 5,
    "M": 6,
    "C": 7,
    "K": 8,
}
NEGATIVE_SERIES_MARKERS = ["УЦЕН", "СОВМЕСТ", "СОВМ", "COMPAT", "COMPATIBLE", "CACTUS", "КОНТРАКТ", "REFURB", "ВОССТ", "REMAN"]

SUBSTITUTE_NEGATIVE_MARKERS = [
    "СОВМЕСТ", "СОВМ", "COMPAT", "COMPATIBLE", "CACTUS",
    "STATIC CONTROL", "PROFILINE", "NV PRINT", "KATUN", "SAKURA", "MYTONE", "MYTONER",
    "REMAN", "REFURB", "ВОССТ", "КОНТРАКТ", "Б/У", "БУ ", " USED ",
    "COPYRITE", "CET", "G&G", "ELP", "GG-", "NV-", "STATICCONTROL",
    "UNIVERSAL", "СТАНДАРТ", "STANDART", "STANDARD", "BLACK&WHITE", "B&W",
    "AQC-", "HCOL-", "HST-", "XST-", "LI-", "STA-", "BULAT", "COLORING",
    "АНАЛОГ", "ANALOG", "АНАЛ", "СОВМЕСТИМЫЙ", "COMPATIBLE TONER", "СОВМЕСТИМ", "NONAME"
]
BAD_OFFER_MARKERS = [
    "УЦЕН", "УЦЕНКА", "РАСПРОДАЖ", "ЛИКВИД", "SALE", "DISCOUNT", "OUTLET",
    "OPENBOX", "OPEN BOX", "OPEN-BOX", "ВИТРИН", "ДЕМО", "DEMO", "DISPLAY",
    "ПОВРЕЖД", "МЯТАЯУПАК", "МЯТАЯ УПАК", "БЕЗУПАК", "БЕЗ УПАК", "ПЕРЕУПАК",
    "REPACK", "REFURB", "REMAN", "ВОССТ", "USED", "Б/У", "БУ ", "УПАКОВКАПОВРЕЖДЕНА"
]
QUALITY_FLAG_COLUMN_MARKERS = [
    "УЦЕН", "УЦЕНКА", "РАСПРОДАЖ", "НЕКОНД", "НЕКОНДИЦ", "ЛИКВИД",
    "SALE", "DISCOUNT", "OUTLET", "OPEN BOX", "OPENBOX", "OPEN-BOX",
    "DEMO", "DISPLAY", "ВИТРИН", "ПОВРЕЖД", "МЯТАЯ", "REPACK",
    "REFURB", "REMAN", "USED", "Б/У", "БУ", "СТОК", "OUT OF BOX"
]
ALL_NEGATIVE_DIST_MARKERS = sorted(set(SUBSTITUTE_NEGATIVE_MARKERS + BAD_OFFER_MARKERS))

SUSPECT_VENDOR_ARTICLE_PREFIX_RE = re.compile(
    r"^(?:MT|GG|CS|ELP|OPC|PCR|WB|CH|SR|LI|HST|XST|HCOL|AQC|STA|NV|SC|BULAT|CET|KATUN|SAKURA|PROFILINE|STATIC)[-/]",
    re.IGNORECASE,
)
POSITIVE_ORIGINAL_MARKERS = ["ОРИГИН", "ORIGINAL", "GENUINE", "OEM", "RETURN PROGRAM"]

RESOURCE_ALLOWED_PRODUCT_TYPES = {"РАСХОДНЫЕ МАТЕРИАЛЫ"}
OCS_ALLOWED_PRODUCT_TYPES = {
    "РАСХОДНЫЕ МАТЕРИАЛЫ ДЛЯ МАТРИЧНЫХ ПРИНТЕРОВ",
    "РАСХОДНЫЕ МАТЕРИАЛЫ ДЛЯ СТРУЙНЫХ ПРИНТЕРОВ",
    "РАСХОДНЫЕ МАТЕРИАЛЫ ДЛЯ ЛАЗЕРНЫХ ПРИНТЕРОВ",
}
MERLION_ALLOWED_GROUP1_TYPES = {"РАСХОДНЫЕ МАТЕРИАЛЫ"}
MERLION_ALLOWED_GROUP2_TYPES = {"ОРИГИНАЛЬНЫЕ"}
MERLION_ALLOWED_PRODUCT_TYPES = {
    "ДРАМ-КАРТРИДЖИ",
    "ЛЕНТОЧНЫЕ КАРТРИДЖИ",
    "НАБОРЫ ДЛЯ ПЕЧАТИ",
    "ПЕЧАТАЮЩИЕ ГОЛОВКИ",
    "СТРУЙНЫЕ КАРТРИДЖИ",
    "ТОНЕР",
    "ТОНЕР-КАРТРИДЖИ",
    "ЧЕРНИЛА",
}
RESOURCE_ALLOWED_BRAND_KEYS = {
    "AVISION", "BROTHER", "CANON", "EPSON", "HP", "KONICAMINOLTA", "KYOCERA",
    "LEXMARK", "OKI", "PANASONIC", "PANTUM", "RICOH", "XEROX", "SAMSUNG",
    "SHARP", "КАТЮША"
}
OCS_ALLOWED_BRAND_KEYS = set(RESOURCE_ALLOWED_BRAND_KEYS)
MERLION_ALLOWED_BRAND_KEYS = set(RESOURCE_ALLOWED_BRAND_KEYS)
RESOURCE_BRAND_KEY_ALIASES = {
    "HEWLETTPACKARD": "HP",
    "HEWLETTPACKARDINC": "HP",
    "HPINC": "HP",
    "KONICAMINOLTA": "KONICAMINOLTA",
    "KONICAMINOLTAINC": "KONICAMINOLTA",
    "XEROXCORPORATION": "XEROX",
}


def has_suspect_vendor_article_prefix(value: object) -> bool:
    raw = normalize_text(value).upper()
    if not raw:
        return False
    return bool(SUSPECT_VENDOR_ARTICLE_PREFIX_RE.match(raw))


def confident_dist_code_count(row: pd.Series) -> int:
    codes = row.get("name_code_list", []) or []
    seen: set[str] = set()
    count = 0
    for code in codes:
        norm = normalize_article(code)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        count += 1
    return count


def is_confident_alt_exact_match(row: pd.Series, token_norm: str) -> bool:
    if not bool(row.get("is_good_offer", True)):
        return False
    if not bool(row.get("is_original", True)):
        return False

    article_raw = row.get("article", "")
    alt_raw = row.get("alt_article", "")
    name_raw = row.get("name", "")
    brand_raw = row.get("brand", "")

    if has_suspect_vendor_article_prefix(article_raw) or has_suspect_vendor_article_prefix(alt_raw):
        return False

    compact_article = compact_text(article_raw)
    compact_alt = compact_text(alt_raw)
    token_compact = compact_text(token_norm)

    if token_compact and compact_article and compact_article != token_compact and token_compact in compact_article:
        return False

    if text_has_any_marker(" ".join([str(article_raw), str(alt_raw), str(name_raw), str(brand_raw)]), SUBSTITUTE_NEGATIVE_MARKERS):
        return False

    code_count = confident_dist_code_count(row)
    if code_count > 4:
        return False

    dist_family = detect_supply_family(article_raw, alt_raw, name_raw)
    if dist_family == "OTHER" and not text_has_any_marker(name_raw, POSITIVE_ORIGINAL_MARKERS):
        return False

    return True


def _pantum_brand_key(value: object) -> str:
    return canonical_brand_key(value)


def pantum_safe_p_alias_match(token_norm: str, row: pd.Series, own_brand: object = "") -> bool:
    token_norm = normalize_article(token_norm)
    if not token_norm:
        return False
    row_brand_key = _pantum_brand_key(row.get("brand", ""))
    own_brand_key = _pantum_brand_key(own_brand)
    if row_brand_key != "PANTUM" and own_brand_key != "PANTUM":
        return False
    if not bool(row.get("is_good_offer", True)) or not bool(row.get("is_original", True)):
        return False
    if has_suspect_vendor_article_prefix(row.get("article", "")) or has_suspect_vendor_article_prefix(row.get("alt_article", "")):
        return False
    if text_has_any_marker(
        " ".join([
            str(row.get("article", "")),
            str(row.get("alt_article", "")),
            str(row.get("name", "")),
            str(row.get("brand", "")),
        ]),
        SUBSTITUTE_NEGATIVE_MARKERS,
    ):
        return False
    if confident_dist_code_count(row) > 4:
        return False

    candidate_codes = [normalize_article(row.get("article", "")), normalize_article(row.get("alt_article", ""))]
    candidate_codes = [code for code in candidate_codes if code]
    if not candidate_codes:
        return False

    safe_pairs = {token_norm}
    if token_norm.endswith("P") and len(token_norm) > 4:
        safe_pairs.add(token_norm[:-1])
    else:
        safe_pairs.add(token_norm + "P")

    for code in candidate_codes:
        if code in safe_pairs:
            return True
    return False


def is_confident_distributor_row_for_choice(row: pd.Series, choice: dict[str, Any], token_norm: str, own_codes: Optional[list[str]] = None) -> bool:
    if not bool(row.get("is_good_offer", True)):
        return False
    if not family_compatible(choice, row):
        return False

    own_brand = choice.get("brand", "")
    code_pool = set(unique_norm_codes((own_codes or []) + row_catalog_compare_codes(choice, token_norm)))
    if not code_pool:
        return False

    row_article_norm = normalize_article(row.get("article", ""))
    row_alt_norm = normalize_article(row.get("alt_article", ""))

    if row_article_norm in code_pool:
        return True
    if row_alt_norm in code_pool:
        return is_confident_alt_exact_match(row, row_alt_norm)
    for code in code_pool:
        if pantum_safe_p_alias_match(code, row, own_brand=own_brand):
            return True
    return False


def init_state() -> None:
    defaults = {
        "catalog_base_df": None,
        "catalog_df": None,
        "catalog_name": "ещё не загружен",
        "article_ref_df": None,
        "article_ref_name": "ещё не загружен",
        "avito_df": None,
        "avito_name": "ещё не загружен",
        "resource_df": None,
        "resource_name": "ещё не загружен",
        "ocs_df": None,
        "ocs_name": "ещё не загружен",
        "merlion_df": None,
        "merlion_name": "ещё не загружен",
        "search_input": "",
        "submitted_query": "",
        "last_result": None,
        "price_mode": "-12%",
        "custom_discount": 10.0,
        "round100": True,
        "search_mode": "Только артикул",
        "template1_footer": DEFAULT_TEMPLATE1_FOOTER,
        "price_patch_input": "",
        "patch_message": "",
        "series_mode": "Только оригиналы",
        "distributor_threshold": 20.0,
        "distributor_min_qty": 1.0,
        "distributor_report_df": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


init_state()


# ------------------------------
# Базовые функции текущего файла
# ------------------------------

def normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = re.sub(r"\s+", " ", str(value).strip())
    if text.lower() in {"nan", "nat", "none"}:
        return ""
    return text


def normalize_article(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = text.translate(CYRILLIC_ARTICLE_TRANSLATION)
    return re.sub(r"[^A-Za-z0-9]", "", text).upper()


def contains_text(value: object) -> str:
    return normalize_text(value).upper()


def compact_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "nat", "none"}:
        return ""
    return re.sub(r"\s+", "", text).upper()


def canonical_brand_key(value: object) -> str:
    raw = contains_text(value)
    if not raw:
        return ""
    key = re.sub(r"[^A-ZА-Я0-9]", "", raw)
    return RESOURCE_BRAND_KEY_ALIASES.get(key, key)


CATALOG_BRAND_PATTERNS: list[tuple[str, str]] = [
    ("KONICA-MINOLTA", "Konica-Minolta"),
    ("KONICA MINOLTA", "Konica-Minolta"),
    ("КАТЮША", "Катюша"),
    ("KYOCERA", "Kyocera"),
    ("LEXMARK", "Lexmark"),
    ("PANASONIC", "Panasonic"),
    ("BROTHER", "Brother"),
    ("CANON", "Canon"),
    ("EPSON", "Epson"),
    ("PANTUM", "Pantum"),
    ("XEROX", "Xerox"),
    ("SAMSUNG", "Samsung"),
    ("SHARP", "Sharp"),
    ("AVISION", "Avision"),
    ("RICOH", "Ricoh"),
    ("OKI", "OKI"),
    ("HP", "HP"),
]

def infer_brand_from_name(name: object) -> str:
    text = contains_text(name)
    if not text:
        return ""
    for needle, label in CATALOG_BRAND_PATTERNS:
        if needle in text:
            return label
    return ""

def normalize_or_infer_brand(raw_brand: object, name: object = "") -> str:
    brand = normalize_text(raw_brand)
    if brand:
        return brand
    return infer_brand_from_name(name)


def first_existing_series(df: pd.DataFrame, candidates: list[str], default: object = "") -> pd.Series:
    for candidate in candidates:
        if candidate in df.columns:
            return df[candidate]
    return pd.Series([default] * len(df), index=df.index)


def is_resource_allowed_type(value: object) -> bool:
    return contains_text(value) in RESOURCE_ALLOWED_PRODUCT_TYPES


def is_resource_allowed_brand(value: object) -> bool:
    key = canonical_brand_key(value)
    return bool(key) and key in RESOURCE_ALLOWED_BRAND_KEYS


def is_ocs_allowed_type(value: object) -> bool:
    return contains_text(value) in OCS_ALLOWED_PRODUCT_TYPES


def is_ocs_allowed_brand(value: object) -> bool:
    key = canonical_brand_key(value)
    return bool(key) and key in OCS_ALLOWED_BRAND_KEYS


def is_merlion_allowed_root(value: object) -> bool:
    return contains_text(value) in MERLION_ALLOWED_GROUP1_TYPES


def is_merlion_allowed_group2(value: object) -> bool:
    return contains_text(value) in MERLION_ALLOWED_GROUP2_TYPES


def is_merlion_allowed_type(value: object) -> bool:
    return contains_text(value) in MERLION_ALLOWED_PRODUCT_TYPES


def is_merlion_allowed_brand(value: object) -> bool:
    key = canonical_brand_key(value)
    return bool(key) and key in MERLION_ALLOWED_BRAND_KEYS


def resource_brand_filter(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df.iloc[0:0].copy()
    out = df.copy()
    if "product_type" in out.columns:
        out = out[out["product_type"].apply(is_resource_allowed_type)].copy()
    if "brand" in out.columns:
        out = out[out["brand"].apply(is_resource_allowed_brand)].copy()
    return out.reset_index(drop=True)


def ocs_brand_filter(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df.iloc[0:0].copy()
    out = df.copy()
    if "product_type" in out.columns:
        out = out[out["product_type"].apply(is_ocs_allowed_type)].copy()
    if "brand" in out.columns:
        out = out[out["brand"].apply(is_ocs_allowed_brand)].copy()
    return out.reset_index(drop=True)


def merlion_brand_filter(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df.iloc[0:0].copy()
    out = df.copy()
    if "group_root" in out.columns:
        out = out[out["group_root"].apply(is_merlion_allowed_root)].copy()
    if "group_level2" in out.columns:
        out = out[out["group_level2"].apply(is_merlion_allowed_group2)].copy()
    if "product_type" in out.columns:
        out = out[out["product_type"].apply(is_merlion_allowed_type)].copy()
    if "brand" in out.columns:
        out = out[out["brand"].apply(is_merlion_allowed_brand)].copy()
    return out.reset_index(drop=True)


def is_candidate_article_norm(norm: str) -> bool:
    if not norm:
        return False
    if norm.isdigit():
        return len(norm) >= 5
    return len(norm) >= 3 and any(ch.isdigit() for ch in norm) and any(ch.isalpha() for ch in norm)


def extract_article_candidates_from_text(text: object) -> list[str]:
    raw = str(text or "").upper()
    prepared = re.sub(r"[|/\\,;:()\[\]{}]+", " ", raw)
    prepared = prepared.replace("№", " ")
    chunks = re.findall(r"[A-ZА-Я0-9._-]{3,}", prepared)
    out: list[str] = []
    seen: set[str] = set()
    for chunk in chunks:
        norm = normalize_article(chunk)
        if not is_candidate_article_norm(norm) or norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
    return out




def unique_norm_codes(items: list[object]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        norm = normalize_article(item)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        out.append(norm)
    return out


def build_catalog_code_list(article: object, name: object) -> list[str]:
    return unique_norm_codes([article, *extract_article_candidates_from_text(name)])


def row_catalog_search_codes(row: pd.Series | dict[str, Any]) -> list[str]:
    existing = row.get("all_code_list", []) or []
    if isinstance(existing, list) and existing:
        return unique_norm_codes(existing)
    return build_catalog_code_list(row.get("article", ""), row.get("name", ""))


def row_catalog_compare_codes(row: pd.Series | dict[str, Any], token: str = "") -> list[str]:
    article = row.get("article", "")
    name = row.get("name", "")
    brand = row.get("brand", "")
    token_norm = normalize_article(token)
    if is_negative_substitute_text(article, name, brand):
        return unique_norm_codes([article, token_norm])
    return unique_norm_codes([token_norm, *row_catalog_search_codes(row)])


def row_has_negative_series_markers(row: pd.Series) -> bool:
    text = f"{row.get('article', '')} {row.get('name', '')}".upper()
    return any(marker in text for marker in NEGATIVE_SERIES_MARKERS)


def split_article_family_suffix(article_norm: str) -> tuple[str, str]:
    m = re.match(r"^(.*?\d)([A-ZА-Я]{1,3})$", article_norm)
    if m:
        return m.group(1), m.group(2)
    return article_norm, ""


def natural_chunks(value: str) -> list[object]:
    parts = re.split(r"(\d+)", value)
    result: list[object] = []
    for part in parts:
        if not part:
            continue
        result.append(int(part) if part.isdigit() else part)
    return result


def series_sort_key(candidate: dict[str, object]) -> tuple[object, ...]:
    article_norm = str(candidate.get("article_norm", ""))
    family, suffix = split_article_family_suffix(article_norm)
    rank = SERIES_SUFFIX_ORDER.get(suffix, 50)
    return (*natural_chunks(family), rank, suffix, article_norm)


def tokenize_text(value: object) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [t for t in re.split(r"[^A-Za-zА-Яа-я0-9]+", text.upper()) if t]


def unique_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        key = normalize_text(item)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def is_negative_substitute_row(row: pd.Series) -> bool:
    text = f"{row.get('article', '')} {row.get('name', '')}".upper()
    markers = [
        "УЦЕН", "СОВМЕСТ", "СОВМ", "COMPAT", "COMPATIBLE", "CACTUS",
        "STATIC CONTROL", "PROFILINE", "NV PRINT", "KATUN", "SAKURA",
        "REMAN", "REFURB", "ВОССТ", "КОНТРАКТ", "Б/У", "БУ ", " USED "
    ]
    return any(marker in text for marker in markers)


def find_column(columns: list[str], candidates: list[str]) -> Optional[str]:
    lower_map = {str(col).strip().lower(): col for col in columns}
    for candidate in candidates:
        col = lower_map.get(candidate.strip().lower())
        if col is not None:
            return col
    for candidate in candidates:
        c_low = candidate.strip().lower()
        for original in columns:
            o_low = str(original).strip().lower()
            if c_low in o_low or o_low in c_low:
                return original
    return None


def detect_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    return {key: find_column(list(df.columns), aliases) for key, aliases in COLUMN_ALIASES.items()}



def detect_article_reference_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    return {key: find_column(list(df.columns), aliases) for key, aliases in ARTICLE_REFERENCE_COLUMN_ALIASES.items()}


@st.cache_data(show_spinner=False)
def load_article_reference_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    bio = io.BytesIO(file_bytes)
    if suffix == ".csv":
        try:
            raw = pd.read_csv(bio)
        except UnicodeDecodeError:
            bio.seek(0)
            raw = pd.read_csv(bio, encoding="cp1251")
    else:
        raw = pd.read_excel(bio)

    raw = raw.dropna(how="all")
    mapping = detect_article_reference_columns(raw)
    if not mapping.get("article") and not mapping.get("manufacturer_article"):
        raise ValueError("Не удалось определить колонки справочника: нужен хотя бы 'Артикул' или 'Артикул производителя'.")
    if not mapping.get("name"):
        raise ValueError("Не удалось определить колонку 'Номенклатура' в справочнике.")

    data = pd.DataFrame()
    data["article"] = raw[mapping["article"]].map(normalize_text) if mapping.get("article") else ""
    data["article_norm"] = raw[mapping["article"]].map(normalize_article) if mapping.get("article") else ""
    data["manufacturer_article"] = raw[mapping["manufacturer_article"]].map(normalize_text) if mapping.get("manufacturer_article") else ""
    data["manufacturer_article_norm"] = raw[mapping["manufacturer_article"]].map(normalize_article) if mapping.get("manufacturer_article") else ""
    data["name"] = raw[mapping["name"]].map(normalize_text)
    data["brand"] = raw.apply(
        lambda r: normalize_or_infer_brand(r[mapping["brand"]], r[mapping["name"]]) if mapping.get("brand") else infer_brand_from_name(r[mapping["name"]]),
        axis=1,
    )
    data["name_code_list"] = data["name"].map(extract_article_candidates_from_text)
    data["ref_code_list"] = data.apply(
        lambda row: unique_norm_codes([row.get("article", ""), row.get("manufacturer_article", ""), *(row.get("name_code_list", []) or [])]),
        axis=1,
    )
    data["is_negative"] = data.apply(
        lambda row: is_negative_substitute_text(row.get("article", ""), row.get("manufacturer_article", ""), row.get("name", ""), row.get("brand", "")),
        axis=1,
    )
    data = data[data["ref_code_list"].map(lambda x: isinstance(x, list) and len(x) > 0)].copy()
    data = data[data["is_negative"] != True].copy()
    data = data.reset_index(drop=True)
    return data


def build_article_reference_lookup(reference_df: pd.DataFrame) -> dict[str, list[int]]:
    lookup: dict[str, list[int]] = {}
    if reference_df is None or reference_df.empty:
        return lookup
    for idx, row in reference_df.iterrows():
        for code in row.get("ref_code_list", []) or []:
            norm = normalize_article(code)
            if not norm:
                continue
            lookup.setdefault(norm, []).append(int(idx))
    return lookup


def expand_catalog_codes_with_reference(catalog_df: pd.DataFrame | None, reference_df: pd.DataFrame | None) -> pd.DataFrame | None:
    if catalog_df is None:
        return None
    if reference_df is None or reference_df.empty:
        out = catalog_df.copy()
        if "reference_code_list" not in out.columns:
            out["reference_code_list"] = [[] for _ in range(len(out))]
        return out

    ref_lookup = build_article_reference_lookup(reference_df)
    out = catalog_df.copy()

    def _expand_row(row: pd.Series) -> pd.Series:
        base_codes = unique_norm_codes(row.get("all_code_list", []) or [row.get("article", ""), *(row.get("name_code_list", []) or [])])
        if not base_codes:
            row["reference_code_list"] = []
            row["all_code_list"] = base_codes
            return row
        if is_negative_substitute_text(row.get("article", ""), row.get("name", ""), row.get("brand", "")):
            row["reference_code_list"] = []
            row["all_code_list"] = base_codes
            return row

        own_brand = str(row.get("brand", "") or "")
        matched_ref_rows: set[int] = set()
        for code in base_codes:
            for ref_idx in ref_lookup.get(code, []):
                ref_row = reference_df.iloc[ref_idx]
                ref_brand = str(ref_row.get("brand", "") or "")
                if own_brand and ref_brand and not brand_match(own_brand, ref_brand):
                    continue
                matched_ref_rows.add(int(ref_idx))

        ref_codes: list[str] = []
        for ref_idx in sorted(matched_ref_rows):
            ref_row = reference_df.iloc[ref_idx]
            ref_codes.extend(ref_row.get("ref_code_list", []) or [])

        row["reference_code_list"] = unique_norm_codes(ref_codes)
        row["all_code_list"] = unique_norm_codes([*base_codes, *row["reference_code_list"]])
        return row

    out = out.apply(_expand_row, axis=1)
    return out


def rebuild_catalog_effective_df() -> None:
    base_df = st.session_state.get("catalog_base_df")
    ref_df = st.session_state.get("article_ref_df")
    if isinstance(base_df, pd.DataFrame):
        st.session_state.catalog_df = expand_catalog_codes_with_reference(base_df, ref_df)
    else:
        st.session_state.catalog_df = None

@st.cache_data(show_spinner=False)
def load_price_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    bio = io.BytesIO(file_bytes)
    if suffix == ".csv":
        try:
            raw = pd.read_csv(bio)
        except UnicodeDecodeError:
            bio.seek(0)
            raw = pd.read_csv(bio, encoding="cp1251")
    else:
        raw = pd.read_excel(bio)

    raw = raw.dropna(how="all")
    mapping = detect_columns(raw)
    required = ["article", "name", "price"]
    missing = [k for k in required if not mapping.get(k)]
    if missing:
        raise ValueError("Не удалось определить обязательные колонки: " + ", ".join(missing))

    data = pd.DataFrame()
    data["article"] = raw[mapping["article"]].map(normalize_text)
    data["article_norm"] = raw[mapping["article"]].map(normalize_article)
    data["name"] = raw[mapping["name"]].map(normalize_text)
    data["brand"] = raw.apply(lambda r: normalize_or_infer_brand(r[mapping["brand"]], r[mapping["name"]]) if mapping.get("brand") else infer_brand_from_name(r[mapping["name"]]), axis=1)
    data["free_qty"] = (
        pd.to_numeric(raw[mapping["free_qty"]], errors="coerce").fillna(0)
        if mapping.get("free_qty")
        else 0
    )
    data["total_qty"] = (
        pd.to_numeric(raw[mapping["total_qty"]], errors="coerce").fillna(0)
        if mapping.get("total_qty")
        else 0
    )
    data["sale_price"] = pd.to_numeric(raw[mapping["price"]], errors="coerce")
    data = data.dropna(subset=["sale_price"])
    data = data[data["article_norm"] != ""].copy()
    data = data.drop_duplicates(subset=["article_norm"], keep="first")

    data["sale_price"] = data["sale_price"].astype(float)
    data["price_12"] = data["sale_price"] * (1 - DEFAULT_DISCOUNT_1 / 100)
    data["price_20"] = data["sale_price"] * (1 - DEFAULT_DISCOUNT_2 / 100)
    data["name_tokens"] = data["name"].map(tokenize_text)
    data["name_code_list"] = data["name"].map(extract_article_candidates_from_text)
    data["all_code_list"] = data.apply(lambda row: build_catalog_code_list(row["article"], row["name"]), axis=1)
    data["search_blob"] = (
        data["article_norm"].fillna("")
        + " "
        + data["name"].fillna("")
        + " "
        + data["brand"].fillna("")
    ).str.upper()
    return data.reset_index(drop=True)


def parse_excel_hyperlink_formula(value: object) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text.startswith("="):
        return "", ""
    m = re.match(r'^=\s*(?:HYPERLINK|ГИПЕРССЫЛКА)\(\s*"([^"]+)"\s*[;,]\s*"([^"]*)"\s*\)$', text, flags=re.IGNORECASE)
    if not m:
        return "", ""
    return m.group(1).strip(), m.group(2).strip()


def cell_display_and_url(cell) -> tuple[str, str]:
    url = ""
    display = ""
    if cell is None:
        return display, url
    try:
        if getattr(cell, "hyperlink", None):
            url = str(cell.hyperlink.target or "").strip()
    except Exception:
        pass
    formula_url, formula_display = parse_excel_hyperlink_formula(cell.value)
    if formula_url:
        url = formula_url
        display = formula_display
    else:
        display = normalize_text(cell.value)
    return display, url


@st.cache_data(show_spinner=False)
def load_avito_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()

    if suffix == ".csv":
        bio = io.BytesIO(file_bytes)
        try:
            raw = pd.read_csv(bio)
        except UnicodeDecodeError:
            bio.seek(0)
            raw = pd.read_csv(bio, encoding="cp1251")

        mapping = {key: find_column(list(raw.columns), aliases) for key, aliases in AVITO_COLUMN_ALIASES.items()}
        if not mapping.get("title"):
            raise ValueError("Не удалось определить колонку 'Название объявления' в файле Авито.")
        rows = []
        for _, r in raw.iterrows():
            ad_id = normalize_text(r[mapping["ad_id"]]) if mapping.get("ad_id") else ""
            title = normalize_text(r[mapping["title"]]) if mapping.get("title") else ""
            url = normalize_text(r[mapping["url"]]) if mapping.get("url") else ""
            price = normalize_text(r[mapping["price"]]) if mapping.get("price") else ""
            if not ad_id and not title:
                continue
            rows.append({
                "ad_id": ad_id,
                "title": title,
                "price": price,
                "url": url,
                "title_codes": extract_article_candidates_from_text(title),
                "title_norm": normalize_text(title).upper(),
            })
        return pd.DataFrame(rows)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws = wb.active
    headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    def find_header_index(candidates: list[str]) -> Optional[int]:
        for idx, header in enumerate(headers, start=1):
            for cand in candidates:
                if header.lower() == cand.lower():
                    return idx
        for idx, header in enumerate(headers, start=1):
            h = header.lower()
            for cand in candidates:
                c = cand.lower()
                if c in h or h in c:
                    return idx
        return None

    ad_id_col = find_header_index(AVITO_COLUMN_ALIASES["ad_id"])
    title_col = find_header_index(AVITO_COLUMN_ALIASES["title"])
    price_col = find_header_index(AVITO_COLUMN_ALIASES["price"])
    url_col = find_header_index(AVITO_COLUMN_ALIASES["url"])

    if not title_col:
        raise ValueError("Не удалось определить колонку 'Название объявления' в файле Авито.")

    rows: list[dict[str, object]] = []
    for r in range(2, ws.max_row + 1):
        ad_display, ad_url = cell_display_and_url(ws.cell(r, ad_id_col)) if ad_id_col else ("", "")
        title_display, title_url = cell_display_and_url(ws.cell(r, title_col))
        explicit_url = normalize_text(ws.cell(r, url_col).value) if url_col else ""
        price_value = normalize_text(ws.cell(r, price_col).value) if price_col else ""
        final_url = explicit_url or title_url or ad_url
        if not ad_display and not title_display:
            continue
        rows.append({
            "ad_id": ad_display,
            "title": title_display,
            "price": price_value,
            "url": final_url,
            "title_codes": extract_article_candidates_from_text(title_display),
            "title_norm": normalize_text(title_display).upper(),
        })
    return pd.DataFrame(rows)


def round_up_to_100(value: float) -> int:
    return int(math.ceil(float(value) / 100.0) * 100)


def round_to_nearest_100(value: float) -> int:
    return int(math.floor(float(value) / 100.0 + 0.5) * 100)


def current_discount(price_mode: str, custom_discount: float) -> float:
    if price_mode == "-12%":
        return DEFAULT_DISCOUNT_1
    if price_mode == "-20%":
        return DEFAULT_DISCOUNT_2
    return max(0.0, float(custom_discount))


def current_price_label(price_mode: str, custom_discount: float) -> str:
    disc = current_discount(price_mode, custom_discount)
    if float(disc).is_integer():
        return f"Цена -{int(disc)}%"
    return f"Цена -{str(round(disc, 2)).replace('.', ',')}%"


def get_selected_price_raw(row: pd.Series, price_mode: str, round100: bool, custom_discount: float) -> float:
    disc = current_discount(price_mode, custom_discount)
    value = float(row["sale_price"]) * (1 - disc / 100)
    return float(round_up_to_100(value)) if round100 else float(round(value, 2))


def fmt_price(value: float | int) -> str:
    if pd.isna(value):
        return ""
    value = float(value)
    if value.is_integer():
        return f"{int(value):,}".replace(",", " ")
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def fmt_price_with_rub(value: float | int) -> str:
    return f"{fmt_price(value)} руб."


def fmt_qty(value: float | int) -> str:
    try:
        v = float(value)
    except Exception:
        return str(value)
    return str(int(v)) if v.is_integer() else f"{v:,.2f}".replace(",", " ").replace(".", ",")


def looks_like_article_token(token: str) -> bool:
    token = normalize_text(token)
    if not token:
        return False
    compact = re.sub(r"[\s\-_./]", "", token)
    has_digit = any(ch.isdigit() for ch in compact)
    has_alpha = any(ch.isalpha() for ch in compact)
    return len(compact) >= 3 and has_digit and has_alpha


def split_query_parts(query: str) -> list[str]:
    parts: list[str] = []
    raw_chunks = re.split(r"[\n,;]+", query)
    for chunk in raw_chunks:
        chunk = normalize_text(chunk)
        if not chunk:
            continue
        if "/" in chunk:
            slash_parts = [normalize_text(x) for x in re.split(r"\s*/\s*", chunk) if normalize_text(x)]
            if len(slash_parts) > 1:
                parts.extend(slash_parts)
                continue
        space_parts = [normalize_text(x) for x in re.split(r"\s+", chunk) if normalize_text(x)]
        if len(space_parts) > 1 and all(looks_like_article_token(x) for x in space_parts):
            parts.extend(space_parts)
        else:
            parts.append(chunk)
    return parts


def normalize_query_for_display(query: str) -> str:
    return "\n".join(split_query_parts(query))


def detect_color(name: str) -> str:
    low = normalize_text(name).lower()
    for needle, label in COLOR_KEYWORDS:
        if needle in low:
            return label
    return ""


def is_available(row: pd.Series) -> bool:
    try:
        return float(row.get("free_qty", 0)) > 0
    except Exception:
        return False


def parse_price_updates(text: str) -> list[tuple[str, float]]:
    updates: list[tuple[str, float]] = []
    for line in text.splitlines():
        line = normalize_text(line)
        if not line:
            continue
        cleaned = line.replace("🔽", " ").replace("🔼", " ").replace("—", "-")
        m = re.search(r"([A-Za-zА-Яа-я0-9./_-]+)\s*-?\s*([0-9][0-9\s.,]*)", cleaned)
        if not m:
            continue
        article = normalize_article(m.group(1))
        price_txt = re.sub(r"[^0-9,\.]", "", m.group(2)).replace(",", ".")
        try:
            price = float(price_txt)
        except ValueError:
            continue
        if article:
            updates.append((article, price))
    return updates


def apply_price_updates(df: pd.DataFrame, updates_text: str) -> tuple[pd.DataFrame, str]:
    updates = parse_price_updates(updates_text)
    if not updates:
        return df, "Не нашёл строк для правки цен."

    out = df.copy()
    updated_count = 0
    missed: list[str] = []
    linked_hits: list[str] = []
    seen_done: set[str] = set()

    for article_norm, new_price in updates:
        if article_norm in seen_done:
            continue
        mask = out["article_norm"] == article_norm
        match_source = "exact"
        if not mask.any():
            linked = out[out["all_code_list"].apply(lambda codes: article_norm in codes if isinstance(codes, list) else False)]
            if not linked.empty:
                safe_linked = linked[~linked.apply(is_negative_substitute_row, axis=1)]
                if not safe_linked.empty:
                    chosen = safe_linked.iloc[0]
                    mask = out["article_norm"] == str(chosen["article_norm"])
                    match_source = "linked"
        if mask.any():
            out.loc[mask, "sale_price"] = float(new_price)
            out.loc[mask, "price_12"] = float(new_price) * (1 - DEFAULT_DISCOUNT_1 / 100)
            out.loc[mask, "price_20"] = float(new_price) * (1 - DEFAULT_DISCOUNT_2 / 100)
            updated_count += 1
            seen_done.add(article_norm)
            if match_source == "linked":
                first_row = out.loc[mask].iloc[0]
                linked_hits.append(f"{article_norm}→{first_row['article']}")
        else:
            missed.append(article_norm)
    message = f"Обновлено цен: {updated_count}"
    if linked_hits:
        message += " | Связанные: " + ", ".join(linked_hits[:8])
    if missed:
        message += " | Не найдено: " + ", ".join(missed[:10])
    return out, message


def find_best_row_for_token(df: pd.DataFrame, token: str, search_mode: str) -> tuple[Optional[pd.Series], str]:
    article_norm = normalize_article(token)
    if not article_norm:
        return None, ""
    exact = df[df["article_norm"] == article_norm]
    if not exact.empty:
        exact_safe = exact[~exact.apply(is_negative_substitute_row, axis=1)]
        chosen = exact_safe.iloc[0] if not exact_safe.empty else exact.iloc[0]
        return chosen, "exact"
    alias_matches = df[df["all_code_list"].apply(lambda codes: article_norm in codes if isinstance(codes, list) else False)]
    if not alias_matches.empty:
        safe_alias_matches = alias_matches[~alias_matches.apply(is_negative_substitute_row, axis=1)]
        chosen = safe_alias_matches.iloc[0] if not safe_alias_matches.empty else alias_matches.iloc[0]
        return chosen, "linked"
    if search_mode in {"Умный", "Артикул + название + бренд"}:
        contains = df[df["search_blob"].str.contains(re.escape(token.upper()), na=False)]
        if not contains.empty:
            safe_contains = contains[~contains.apply(is_negative_substitute_row, axis=1)]
            if not safe_contains.empty:
                chosen = safe_contains.iloc[0]
                return chosen, "similar"
            return None, ""
    return None, ""


def resolve_query_tokens(df: pd.DataFrame, query: str, search_mode: str) -> tuple[list[tuple[str, pd.Series, str]], list[str]]:
    resolved: list[tuple[str, pd.Series, str]] = []
    missing: list[str] = []
    for part in split_query_parts(query):
        row, match_type = find_best_row_for_token(df, part, search_mode)
        if row is None:
            missing.append(part)
        else:
            resolved.append((part, row, match_type))
    return resolved, missing


def perform_search(df: pd.DataFrame, query: str, search_mode: str) -> pd.DataFrame:
    resolved, _ = resolve_query_tokens(df, query, search_mode)
    if not resolved:
        return df.iloc[0:0].copy()
    rows = []
    seen_articles: set[str] = set()
    rank_map = {"exact": 0, "linked": 1, "similar": 2}
    for part, row, match_type in resolved:
        article_key = str(row["article_norm"])
        if article_key in seen_articles:
            continue
        seen_articles.add(article_key)
        row_dict = row.to_dict()
        row_dict["match_type"] = match_type
        row_dict["match_query"] = part
        row_dict["_rank"] = rank_map.get(match_type, 99)
        rows.append(row_dict)
    out = pd.DataFrame(rows).sort_values(["_rank", "article_norm"]).drop(columns=["_rank"]).reset_index(drop=True)
    return out


def compact_multiline(text: str) -> str:
    lines = [normalize_text(line) for line in str(text).splitlines()]
    lines = [line for line in lines if line]
    return "\n".join(lines)


def build_offer_template(df: pd.DataFrame, query: str, round100: bool, footer_text: str, search_mode: str) -> str:
    parts = split_query_parts(query)
    if not parts:
        return ""
    groups: dict[str, dict] = {}
    missing_tokens: list[str] = []
    for part in parts:
        row, _ = find_best_row_for_token(df, part, search_mode)
        if row is None:
            missing_tokens.append(part)
            continue
        key = str(row["article_norm"])
        grp = groups.setdefault(key, {"row": row, "tokens": []})
        grp["tokens"].append(part)
    lines: list[str] = []
    hashtag_parts: list[str] = []
    for item in groups.values():
        row = item["row"]
        tokens = unique_preserve_order([normalize_article(t) for t in item["tokens"] if normalize_article(t)])
        if not tokens:
            tokens = [str(row["article_norm"])]
        if is_available(row):
            avito_raw = float(row["sale_price"]) * (1 - DEFAULT_DISCOUNT_1 / 100)
            cash_raw = avito_raw * 0.90
            if round100:
                avito = round_up_to_100(avito_raw)
                cash = round_to_nearest_100(cash_raw)
            else:
                avito = round(avito_raw)
                cash = round(cash_raw)
            head = f"{row['article']} --- {fmt_price_with_rub(avito)} - Авито / {fmt_price_with_rub(cash)} за наличный расчет"
        else:
            color = detect_color(str(row["name"]))
            prefix = f"{row['article']} {color}".strip()
            head = f"{prefix} --- продан"
        lines.append(head)
        hashtag_parts.extend([f"#{t}" for t in tokens])
    for token in missing_tokens:
        lines.append(f"{token} --- продан")
        tok = normalize_article(token)
        if tok:
            hashtag_parts.append(f"#{tok}")
    hashtag_parts = unique_preserve_order(hashtag_parts)
    footer = compact_multiline(footer_text)
    out_lines = [line for line in lines if normalize_text(line)]
    if footer:
        out_lines.extend(footer.split("\n"))
    if hashtag_parts:
        out_lines.append(",".join(hashtag_parts))
    return "\n".join(out_lines)


def build_selected_price_template(df: pd.DataFrame, query: str, price_mode: str, round100: bool, custom_discount: float, search_mode: str) -> str:
    parts = split_query_parts(query)
    lines: list[str] = []
    seen_articles: set[str] = set()
    for part in parts:
        row, _ = find_best_row_for_token(df, part, search_mode)
        if row is None:
            continue
        key = str(row["article_norm"])
        if key in seen_articles or not is_available(row):
            continue
        seen_articles.add(key)
        selected_price = get_selected_price_raw(row, price_mode, round100, custom_discount)
        lines.append(f"{normalize_text(row['name'])} --- {fmt_price_with_rub(selected_price)}")
    return "\n\n".join(lines)


def find_avito_ads(avito_df: pd.DataFrame, query: str, result_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    if avito_df is None or avito_df.empty:
        return pd.DataFrame()
    query_tokens = unique_preserve_order([normalize_article(x) for x in split_query_parts(query) if normalize_article(x)])
    token_pool = list(query_tokens)
    if isinstance(result_df, pd.DataFrame) and not result_df.empty:
        for _, row in result_df.iterrows():
            art = normalize_article(row.get("article"))
            if art:
                token_pool.append(art)
            for code in row.get("name_code_list", []) or []:
                norm = normalize_article(code)
                if norm:
                    token_pool.append(norm)
    token_pool = unique_preserve_order(token_pool)
    if not token_pool:
        return pd.DataFrame()
    matches: list[dict[str, object]] = []
    for _, row in avito_df.iterrows():
        codes = [normalize_article(x) for x in (row.get("title_codes", []) or []) if normalize_article(x)]
        matched_tokens = [tok for tok in token_pool if tok in codes]
        match_kind = ""
        if matched_tokens:
            match_kind = "exact" if any(tok in query_tokens for tok in matched_tokens) else "related"
        else:
            title_norm = str(row.get("title_norm", ""))
            boundary_hits = [tok for tok in token_pool if tok and re.search(rf"(?<![A-ZА-Я0-9]){re.escape(tok)}(?![A-ZА-Я0-9])", title_norm)]
            if boundary_hits:
                matched_tokens = boundary_hits
                match_kind = "exact" if any(tok in query_tokens for tok in matched_tokens) else "related"
        if matched_tokens:
            row_dict = dict(row)
            row_dict["matched_tokens"] = unique_preserve_order(matched_tokens)
            row_dict["match_score"] = len(row_dict["matched_tokens"])
            row_dict["match_kind"] = match_kind or "related"
            matches.append(row_dict)
    if not matches:
        return pd.DataFrame()
    out = pd.DataFrame(matches)
    rank = {"exact": 0, "related": 1}
    out["_rank"] = out["match_kind"].map(lambda x: rank.get(str(x), 99))
    out = out.sort_values(["_rank", "match_score", "ad_id", "title"], ascending=[True, False, True, True]).drop(columns=["_rank"]).reset_index(drop=True)
    return out


# ------------------------------------------
# Новые функции: дистрибьютеры и сравнение
# ------------------------------------------

def text_has_any_marker(raw_text: str, markers: list[str]) -> bool:
    compact = compact_text(raw_text)
    spaced = contains_text(raw_text)
    for marker in markers:
        marker_compact = compact_text(marker)
        marker_spaced = contains_text(marker)
        if marker_compact and marker_compact in compact:
            return True
        if marker_spaced and marker_spaced in spaced:
            return True
    return False


def is_truthy_flag_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    text = contains_text(value)
    compact = compact_text(value)
    if not text and not compact:
        return False
    if compact in {"", "0", "НЕТ", "NO", "FALSE", "NONE", "ОК", "OK", "НОРМА", "НОВЫЙ"}:
        return False
    if text in {"-", "--", "---"}:
        return False
    return True


def is_negative_substitute_text(*parts: Any) -> bool:
    return text_has_any_marker(" ".join(str(p or "") for p in parts), SUBSTITUTE_NEGATIVE_MARKERS)


def is_bad_offer_text(*parts: Any) -> bool:
    return text_has_any_marker(" ".join(str(p or "") for p in parts), BAD_OFFER_MARKERS)


def row_explicitly_flagged_bad(row: pd.Series) -> bool:
    return is_truthy_flag_value(row.get("quality_flags", ""))


def row_has_bad_offer_markers(row: pd.Series) -> bool:
    text = " ".join([
        str(row.get("article", "") or ""),
        str(row.get("alt_article", "") or ""),
        str(row.get("name", "") or ""),
        str(row.get("brand", "") or ""),
        str(row.get("group2", "") or ""),
        str(row.get("quality_flags", "") or ""),
    ])
    return is_bad_offer_text(text)


def collect_quality_flag_text(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype=str)
    selected: list[str] = []
    for col in df.columns:
        col_text = contains_text(col)
        if any(contains_text(marker) in col_text for marker in QUALITY_FLAG_COLUMN_MARKERS):
            selected.append(col)
    if not selected:
        return pd.Series([""] * len(df), index=df.index, dtype=object)
    result = pd.Series([""] * len(df), index=df.index, dtype=object)
    for col in selected:
        result = result.astype(str) + " " + df[col].fillna("").astype(str)
    return result.str.strip()


def parse_resource_qty(value: Any) -> float:
    text = contains_text(value)
    compact = compact_text(value)
    try:
        parsed = float(str(value).replace(" ", "").replace(",", "."))
        return max(0.0, parsed)
    except Exception:
        pass
    if compact in {"+++", "МНОГО"}:
        return 10.0
    if compact in {"++"}:
        return 5.0
    if compact in {"+"}:
        return 1.0
    if any(x in text for x in ["НЕТ", "ПОД ЗАКАЗ", "ОЖИДАЕТСЯ"]):
        return 0.0
    return 0.0


def parse_ocs_qty(value: Any) -> float:
    text = contains_text(value)
    compact = compact_text(value)
    try:
        parsed = float(str(value).replace(" ", "").replace(",", "."))
        return max(0.0, parsed)
    except Exception:
        pass
    if compact in {"ЕСТЬ", "+", "+++"}:
        return 10.0
    if any(marker in text for marker in ["ПОД ЗАКАЗ", "ОЖИДАЕТСЯ", "НЕТ"]):
        return 0.0
    return 0.0


def parse_merlion_qty(value: Any) -> float:
    text = compact_text(value)
    if not text:
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        pass
    if text in {"+++", "МНОГО"}:
        return 10.0
    if text in {"+", "МАЛО"}:
        return 1.0
    return 0.0


def standardize_distributor_result(data: pd.DataFrame, distributor: str) -> pd.DataFrame:
    data = data.copy()
    if "alt_article" not in data.columns:
        data["alt_article"] = ""
    if "alt_article_norm" not in data.columns:
        data["alt_article_norm"] = data["alt_article"].map(normalize_article)
    if "quality_flags" not in data.columns:
        data["quality_flags"] = ""
    if "product_type" not in data.columns:
        data["product_type"] = ""

    data["article"] = data["article"].fillna("").astype(str).map(normalize_text)
    data["article_norm"] = data["article"].map(normalize_article)
    data["alt_article"] = data["alt_article"].fillna("").astype(str).map(normalize_text)
    data["alt_article_norm"] = data["alt_article"].map(normalize_article)
    data["name"] = data["name"].fillna("").astype(str).map(normalize_text)
    data["brand"] = data["brand"].fillna("").astype(str).map(normalize_text)
    data["quality_flags"] = data["quality_flags"].fillna("").astype(str).map(normalize_text)
    data["product_type"] = data["product_type"].fillna("").astype(str).map(normalize_text)
    data["free_qty"] = pd.to_numeric(data["free_qty"], errors="coerce").fillna(0)
    data["price"] = pd.to_numeric(data["price"], errors="coerce").fillna(0)
    data["name_tokens"] = data["name"].map(tokenize_text)
    data["name_code_list"] = data["name"].map(extract_article_candidates_from_text)
    data["search_blob"] = (
        data["article"].astype(str)
        + " " + data["alt_article"].astype(str)
        + " " + data["name"].astype(str)
        + " " + data["brand"].astype(str)
        + " " + data["product_type"].astype(str)
        + " " + data["quality_flags"].astype(str)
    ).map(contains_text)
    data["distributor"] = distributor
    data = data[(data["article_norm"] != "") | (data["alt_article_norm"] != "")].copy()
    data = data[data["price"] > 0].copy()
    return data.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_resource_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Price", header=1)
    df = df.dropna(how="all")
    data = pd.DataFrame()
    data["article"] = df.get("Артикул", "").map(normalize_text)
    data["alt_article"] = df.get("Артикул производителя", "").map(normalize_text)
    data["name"] = df.get("Номенклатура", "").map(normalize_text)
    data["brand"] = df.get("Производитель", "").map(normalize_text)
    data["product_type"] = df.get("Тип продукции", "").map(normalize_text) if "Тип продукции" in df.columns else ""
    data["price"] = pd.to_numeric(df.get("Цена, руб", 0), errors="coerce")
    data["free_qty"] = df.get("Доступно Москва", 0).map(parse_resource_qty)
    data["quality_flags"] = collect_quality_flag_text(df)
    data = standardize_distributor_result(data, "Ресурс")
    data["resource_type_ok"] = data["product_type"].apply(is_resource_allowed_type)
    data["resource_brand_ok"] = data["brand"].apply(is_resource_allowed_brand)
    data["is_original"] = ~data.apply(lambda r: is_negative_substitute_text(r["article"], r["alt_article"], r["name"], r["brand"]), axis=1)
    data["is_good_offer"] = (
        data["resource_type_ok"]
        & data["resource_brand_ok"]
        & data["is_original"]
        & ~data.apply(row_has_bad_offer_markers, axis=1)
        & ~data.apply(row_explicitly_flagged_bad, axis=1)
    )
    data = data[data["resource_type_ok"] & data["resource_brand_ok"]].copy()
    data = data[data["free_qty"] > 0].copy()
    return data.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_ocs_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Наличие и цены")
    df = df.dropna(how="all")
    data = pd.DataFrame()
    data["article"] = df.get("Каталожный номер", "").map(normalize_text)
    data["alt_article"] = df.get("Номенклатурный номер", "").map(normalize_text) if "Номенклатурный номер" in df.columns else ""
    data["name"] = df.get("Наименование", "").map(normalize_text)
    data["brand"] = df.get("Производитель", "").map(normalize_text)
    data["product_type"] = df.get("Категория оборудования", "").map(normalize_text) if "Категория оборудования" in df.columns else ""
    data["price"] = pd.to_numeric(df.get("Цена", 0), errors="coerce")
    data["free_qty"] = df.get("Доступно для резерва", 0).map(parse_ocs_qty)
    data["quality_flags"] = collect_quality_flag_text(df)
    data = standardize_distributor_result(data, "OCS")
    data["ocs_type_ok"] = data["product_type"].apply(is_ocs_allowed_type)
    data["ocs_brand_ok"] = data["brand"].apply(is_ocs_allowed_brand)
    data["is_original"] = ~data.apply(lambda r: is_negative_substitute_text(r["article"], r["alt_article"], r["name"], r["brand"]), axis=1)
    data["is_good_offer"] = (
        data["ocs_type_ok"]
        & data["ocs_brand_ok"]
        & data["is_original"]
        & ~data.apply(row_has_bad_offer_markers, axis=1)
        & ~data.apply(row_explicitly_flagged_bad, axis=1)
    )
    data = data[data["ocs_type_ok"] & data["ocs_brand_ok"]].copy()
    data = data[data["free_qty"] > 0].copy()
    return data.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_merlion_file(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Price List", header=4)
    df = df.dropna(how="all")
    data = pd.DataFrame()
    data["article"] = first_existing_series(df, ["Код производителя"], "").map(normalize_text)
    data["alt_article"] = first_existing_series(df, ["Доп. Номер"], "").map(normalize_text)
    data["name"] = first_existing_series(df, ["Наименование"], "").map(normalize_text)
    data["brand"] = first_existing_series(df, ["Бренд", "Производитель"], "").map(normalize_text)
    data["group_root"] = first_existing_series(df, ["Группа 1", "Группа1", "Товарная группа", "Группа"], "").map(normalize_text)
    data["group_level2"] = first_existing_series(df, ["Группа 2", "Группа2", "Подгруппа", "Категория"], "").map(normalize_text)
    data["product_type"] = first_existing_series(df, ["Группа 3", "Группа3", "Вид товара", "Подкатегория"], "").map(normalize_text)
    data["price"] = pd.to_numeric(first_existing_series(df, ["Цена(руб)", "Цена"], 0), errors="coerce")
    data["free_qty"] = first_existing_series(df, ["Доступно", "Наличие"], 0).map(parse_merlion_qty)
    data["quality_flags"] = collect_quality_flag_text(df)
    data = standardize_distributor_result(data, "Мерлион")
    data["merlion_root_ok"] = data["group_root"].apply(is_merlion_allowed_root)
    data["merlion_group2_ok"] = data["group_level2"].apply(is_merlion_allowed_group2)
    data["merlion_type_ok"] = data["product_type"].apply(is_merlion_allowed_type)
    data["merlion_brand_ok"] = data["brand"].apply(is_merlion_allowed_brand)
    data["is_original"] = (
        data["merlion_root_ok"]
        & data["merlion_group2_ok"]
        & data["merlion_type_ok"]
        & data["merlion_brand_ok"]
        & ~data.apply(lambda r: is_negative_substitute_text(r["article"], r["alt_article"], r["name"], r["brand"], r.get("group_root", ""), r.get("group_level2", ""), r.get("product_type", "")), axis=1)
    )
    data["is_good_offer"] = (
        data["merlion_root_ok"]
        & data["merlion_group2_ok"]
        & data["merlion_type_ok"]
        & data["merlion_brand_ok"]
        & data["is_original"]
        & ~data.apply(row_has_bad_offer_markers, axis=1)
        & ~data.apply(row_explicitly_flagged_bad, axis=1)
    )
    data = data[data["merlion_root_ok"] & data["merlion_group2_ok"] & data["merlion_type_ok"] & data["merlion_brand_ok"]].copy()
    data = data[data["free_qty"] > 0].copy()
    return data.reset_index(drop=True)


def distributor_sources_ready() -> bool:
    for key in ["resource_df", "ocs_df", "merlion_df"]:
        df = st.session_state.get(key)
        if isinstance(df, pd.DataFrame) and not df.empty:
            return True
    return False


def brand_match(catalog_brand: str, dist_brand: str) -> bool:
    a = compact_text(catalog_brand)
    b = compact_text(dist_brand)
    if not a or not b:
        return True
    return a in b or b in a


def detect_supply_family(*parts: Any) -> str:
    text = contains_text(" ".join(str(p or "") for p in parts))
    family_markers = [
        ("CHIP", ["ЧИП", " CHIP "]),
        ("DRUM", ["ФОТОБАРАБ", "БАРАБАН", "DRUM", "OPC", "IMAGING UNIT", "IMAGE UNIT"]),
        ("BLADE", ["РАКЕЛ", "ЛЕЗВИ", "WIPER", "BLADE", "DOCTOR BLADE", "ДОЗИРУЮЩ"]),
        ("DEVELOPER", ["ДЕВЕЛОП", "DEVELOPER"]),
        ("FUSER", ["ПЕЧКА", "FUSER"]),
        ("BELT", ["BELT", "ЛЕНТА ПЕРЕНОСА", "TRANSFER BELT"]),
        ("BOTTLE", ["БУТЫЛ", "BOTTLE", "WASTE TONER"]),
        ("CARTRIDGE", ["КАРТРИДЖ", "TONER CARTRIDGE", "INK CARTRIDGE", "RIBBON", "ТОНЕР", " TONER ", " INK "]),
    ]
    for family, markers in family_markers:
        for marker in markers:
            if contains_text(marker).strip() in text:
                return family
    return "OTHER"


def family_compatible(own_row: dict[str, Any], dist_row: pd.Series) -> bool:
    own_family = detect_supply_family(own_row.get("article", ""), own_row.get("name", ""))
    dist_family = detect_supply_family(dist_row.get("article", ""), dist_row.get("alt_article", ""), dist_row.get("name", ""))
    if own_family == "OTHER":
        return True
    if dist_family == "OTHER":
        return False
    return own_family == dist_family



def resource_search_candidates(df: pd.DataFrame, token_norm: str, own_article_norm: str, search_mode: str, own_codes: Optional[list[str]] = None) -> pd.DataFrame:
    if df is None or df.empty:
        return df.iloc[0:0].copy()

    working = df.copy()
    if "is_good_offer" in working.columns:
        working = working[working["is_good_offer"] == True].copy()
    if working.empty:
        return working
    working = resource_brand_filter(working)
    if working.empty:
        return working

    search_codes = unique_norm_codes([token_norm, own_article_norm, *(own_codes or [])])
    if not search_codes:
        return working.iloc[0:0].copy()

    primary_exact = working[working["article_norm"].isin(search_codes)].copy()
    if not primary_exact.empty:
        primary_exact["_match_rank"] = 0
        return primary_exact

    alt_exact = working[working["alt_article_norm"].isin(search_codes)].copy()
    if not alt_exact.empty:
        alt_exact = alt_exact[alt_exact.apply(lambda r: is_confident_alt_exact_match(r, next((c for c in search_codes if normalize_article(r.get("alt_article", "")) == c), token_norm or own_article_norm)), axis=1)].copy()
        if not alt_exact.empty:
            alt_exact["_match_rank"] = 1
            return alt_exact

    pantum_p = working[working.apply(lambda r: any(pantum_safe_p_alias_match(code, r) for code in search_codes), axis=1)].copy()
    if not pantum_p.empty:
        pantum_p = pantum_p[~pantum_p["article_norm"].isin(search_codes) & ~pantum_p["alt_article_norm"].isin(search_codes)].copy()
        if not pantum_p.empty:
            pantum_p["_match_rank"] = 2
            return pantum_p

    name_code = working[working["name_code_list"].apply(lambda codes: any(code in codes for code in search_codes) if isinstance(codes, list) else False)].copy()
    if not name_code.empty:
        name_code = name_code[name_code.apply(lambda r: is_confident_alt_exact_match(r, token_norm or own_article_norm), axis=1)].copy()
        if not name_code.empty:
            name_code["_match_rank"] = 3
            return name_code

    return working.iloc[0:0].copy()


def ocs_search_candidates(df: pd.DataFrame, token_norm: str, own_article_norm: str, search_mode: str, own_codes: Optional[list[str]] = None) -> pd.DataFrame:
    if df is None or df.empty:
        return df.iloc[0:0].copy()

    working = df.copy()
    if "is_good_offer" in working.columns:
        working = working[working["is_good_offer"] == True].copy()
    if working.empty:
        return working
    working = ocs_brand_filter(working)
    if working.empty:
        return working

    search_codes = unique_norm_codes([token_norm, own_article_norm, *(own_codes or [])])
    if not search_codes:
        return working.iloc[0:0].copy()

    primary_exact = working[working["article_norm"].isin(search_codes)].copy()
    if not primary_exact.empty:
        primary_exact["_match_rank"] = 0
        return primary_exact

    alt_exact = working[working["alt_article_norm"].isin(search_codes)].copy()
    if not alt_exact.empty:
        alt_exact = alt_exact[alt_exact.apply(lambda r: is_confident_alt_exact_match(r, next((c for c in search_codes if normalize_article(r.get("alt_article", "")) == c), token_norm or own_article_norm)), axis=1)].copy()
        if not alt_exact.empty:
            alt_exact["_match_rank"] = 1
            return alt_exact

    pantum_p = working[working.apply(lambda r: any(pantum_safe_p_alias_match(code, r) for code in search_codes), axis=1)].copy()
    if not pantum_p.empty:
        pantum_p = pantum_p[~pantum_p["article_norm"].isin(search_codes) & ~pantum_p["alt_article_norm"].isin(search_codes)].copy()
        if not pantum_p.empty:
            pantum_p["_match_rank"] = 2
            return pantum_p

    name_code = working[working["name_code_list"].apply(lambda codes: any(code in codes for code in search_codes) if isinstance(codes, list) else False)].copy()
    if not name_code.empty:
        name_code = name_code[name_code.apply(lambda r: is_confident_alt_exact_match(r, token_norm or own_article_norm), axis=1)].copy()
        if not name_code.empty:
            name_code["_match_rank"] = 3
            return name_code

    return working.iloc[0:0].copy()


def merlion_search_candidates(df: pd.DataFrame, token_norm: str, own_article_norm: str, search_mode: str, own_codes: Optional[list[str]] = None) -> pd.DataFrame:
    if df is None or df.empty:
        return df.iloc[0:0].copy()

    working = df.copy()
    if "is_good_offer" in working.columns:
        working = working[working["is_good_offer"] == True].copy()
    if working.empty:
        return working
    working = merlion_brand_filter(working)
    if working.empty:
        return working

    search_codes = unique_norm_codes([token_norm, own_article_norm, *(own_codes or [])])
    if not search_codes:
        return working.iloc[0:0].copy()

    alt_exact = working[working["alt_article_norm"].isin(search_codes)].copy()
    if not alt_exact.empty:
        alt_exact = alt_exact[alt_exact.apply(lambda r: is_confident_alt_exact_match(r, next((c for c in search_codes if normalize_article(r.get("alt_article", "")) == c), token_norm or own_article_norm)), axis=1)].copy()
        if not alt_exact.empty:
            alt_exact["_match_rank"] = 0
            return alt_exact

    primary_exact = working[working["article_norm"].isin(search_codes)].copy()
    if not primary_exact.empty:
        primary_exact["_match_rank"] = 1
        return primary_exact

    pantum_p = working[working.apply(lambda r: any(pantum_safe_p_alias_match(code, r) for code in search_codes), axis=1)].copy()
    if not pantum_p.empty:
        pantum_p = pantum_p[~pantum_p["article_norm"].isin(search_codes) & ~pantum_p["alt_article_norm"].isin(search_codes)].copy()
        if not pantum_p.empty:
            pantum_p["_match_rank"] = 2
            return pantum_p

    linked = working[working["name_code_list"].apply(lambda codes: any(code in codes for code in search_codes) if isinstance(codes, list) else False)].copy()
    if not linked.empty:
        linked = linked[linked.apply(lambda r: is_confident_alt_exact_match(r, token_norm or own_article_norm), axis=1)].copy()
        if not linked.empty:
            linked["_match_rank"] = 3
            return linked

    return working.iloc[0:0].copy()


def distributor_search_candidates(df: pd.DataFrame, token_norm: str, own_article_norm: str, search_mode: str, own_codes: Optional[list[str]] = None) -> pd.DataFrame:
    if df is None or df.empty:
        return df.iloc[0:0].copy()

    distributor_name = ""
    try:
        distributor_name = str(df["distributor"].iloc[0])
    except Exception:
        distributor_name = ""
    if distributor_name == "Ресурс":
        return resource_search_candidates(df, token_norm, own_article_norm, search_mode, own_codes=own_codes)
    if distributor_name == "OCS":
        return ocs_search_candidates(df, token_norm, own_article_norm, search_mode, own_codes=own_codes)
    if distributor_name == "Мерлион":
        return merlion_search_candidates(df, token_norm, own_article_norm, search_mode, own_codes=own_codes)

    working = df.copy()
    if "is_good_offer" in working.columns:
        working = working[working["is_good_offer"] == True].copy()
    if working.empty:
        return working

    search_codes = unique_norm_codes([token_norm, own_article_norm, *(own_codes or [])])
    if not search_codes:
        return working.iloc[0:0].copy()

    primary_exact = working[working["article_norm"].isin(search_codes)].copy()
    if not primary_exact.empty:
        primary_exact["_match_rank"] = 0
        return primary_exact

    alt_exact = working[working["alt_article_norm"].isin(search_codes)].copy()
    if not alt_exact.empty:
        alt_exact = alt_exact[alt_exact.apply(lambda r: is_confident_alt_exact_match(r, next((c for c in search_codes if normalize_article(r.get("alt_article", "")) == c), token_norm or own_article_norm)), axis=1)].copy()
        if not alt_exact.empty:
            alt_exact["_match_rank"] = 1
            return alt_exact

    if any(looks_like_article_token(code) for code in search_codes):
        return working.iloc[0:0].copy()

    linked = working[working["name_code_list"].apply(lambda codes: any(code in codes for code in search_codes) if isinstance(codes, list) else False)].copy()
    if not linked.empty:
        linked = linked[linked.apply(lambda r: is_confident_alt_exact_match(r, token_norm or own_article_norm), axis=1)].copy()
        if not linked.empty:
            linked["_match_rank"] = 2
            return linked

    if search_mode != "Только артикул":
        name_contains = working[working["search_blob"].str.contains(re.escape(token_norm), na=False, regex=True)].copy()
        if not name_contains.empty:
            name_contains = name_contains[name_contains.apply(lambda r: is_confident_alt_exact_match(r, token_norm or own_article_norm), axis=1)].copy()
            if not name_contains.empty:
                name_contains["_match_rank"] = 3
                return name_contains

    return working.iloc[0:0].copy()


def get_best_distributor_match_for_source(df: pd.DataFrame, choice: dict[str, Any], token: str, search_mode: str, min_qty: float = 1.0) -> dict[str, Any] | None:
    if df is None or df.empty:
        return None

    own_price = float(choice.get("sale_price", 0) or 0)
    own_brand = str(choice.get("brand", "") or "")
    own_article_norm = normalize_article(choice.get("article", ""))
    token_norm = normalize_article(token)
    own_is_original = not is_negative_substitute_text(choice.get("article", ""), choice.get("name", ""), choice.get("brand", ""))
    own_compare_codes = row_catalog_compare_codes(choice, token)

    cand = distributor_search_candidates(df, token_norm, own_article_norm, search_mode, own_codes=own_compare_codes)
    if cand.empty:
        return None
    cand = cand[cand["free_qty"].astype(float) >= float(min_qty)].copy()
    if "is_good_offer" in cand.columns:
        cand = cand[cand["is_good_offer"] == True].copy()
    if cand.empty:
        return None
    if own_is_original:
        orig = cand[cand["is_original"] == True].copy() if "is_original" in cand.columns else cand.copy()
        if orig.empty:
            return None
        cand = orig
    cand = cand[cand.apply(lambda r: is_confident_distributor_row_for_choice(r, choice, token_norm, own_codes=own_compare_codes), axis=1)].copy()
    if cand.empty:
        return None
    if own_brand:
        brand_filtered = cand[cand["brand"].apply(lambda x: brand_match(own_brand, str(x)))]
        if not brand_filtered.empty:
            cand = brand_filtered
    sort_cols = [c for c in ["_match_rank", "price", "free_qty", "article_norm"] if c in cand.columns]
    ascending = [True, True, False, True][: len(sort_cols)]
    cand = cand.sort_values(sort_cols, ascending=ascending)
    row = cand.iloc[0]
    price = float(row["price"])
    offer = {
        "distributor": str(row.get("distributor", "")),
        "price": price,
        "price_fmt": fmt_price(price),
        "article": str(row.get("article", "")),
        "name": str(row.get("name", "")),
        "brand": str(row.get("brand", "")),
        "free_qty": float(row.get("free_qty", 0) or 0),
        "match_rank": int(row.get("_match_rank", 99) or 99),
    }
    if own_price > 0:
        delta = own_price - price
        delta_percent = (delta / own_price) * 100.0
        offer["delta"] = delta
        offer["delta_fmt"] = fmt_price(abs(delta))
        offer["delta_percent"] = delta_percent
        offer["delta_percent_fmt"] = f"{delta_percent:.1f}".replace(".0", "")
        if abs(delta) < 1e-9:
            offer["status"] = "цена равна"
        elif delta > 0:
            offer["status"] = "лучше нас"
        else:
            offer["status"] = "дороже нас"
    return offer



def get_distributor_offers_for_choice(choice: dict[str, Any], token: str, search_mode: str, min_qty: float = 1.0) -> list[dict[str, Any]]:
    offers: list[dict[str, Any]] = []
    for state_key in ["resource_df", "ocs_df", "merlion_df"]:
        df = st.session_state.get(state_key)
        if df is None or df.empty:
            continue
        offer = get_best_distributor_match_for_source(df, choice, token, search_mode, min_qty=min_qty)
        if offer:
            offers.append(offer)
    offers.sort(key=lambda x: (float(x.get("price", 0) or 0), -float(x.get("free_qty", 0) or 0), str(x.get("distributor", ""))))
    return offers



def find_best_distributor_offer_for_choice(choice: dict[str, Any], token: str, search_mode: str, min_qty: float = 1.0) -> dict[str, Any] | None:
    own_price = float(choice.get("sale_price", 0) or 0)
    best = None
    for offer in get_distributor_offers_for_choice(choice, token, search_mode, min_qty=min_qty):
        price = float(offer.get("price", 0) or 0)
        if own_price > 0 and price >= own_price:
            continue
        if best is None or price < float(best["price"]):
            best = offer
    return best


def build_distributor_compare(result_df: pd.DataFrame, search_mode: str, min_qty: float = 1.0) -> list[dict[str, Any]]:
    if result_df is None or result_df.empty:
        return []
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for _, row in result_df.iterrows():
        row_key = str(row.get("article_norm") or normalize_article(row.get("article", "")))
        if row_key in seen:
            continue
        seen.add(row_key)
        choice = {
            "article": row.get("article", ""),
            "name": row.get("name", ""),
            "brand": row.get("brand", ""),
            "sale_price": float(row.get("sale_price", 0) or 0),
            "row_key": row_key,
        }
        best_offer = find_best_distributor_offer_for_choice(choice, str(row.get("article", "")), search_mode, min_qty=min_qty)
        sale_price = float(row.get("sale_price", 0) or 0)
        out.append({
            "row_key": row_key,
            "article": str(row.get("article", "")),
            "name": str(row.get("name", "")),
            "sale_price": sale_price,
            "sale_price_fmt": fmt_price(sale_price),
            "best_offer": best_offer,
        })
    return out


def distributor_compare_map(result_df: pd.DataFrame, search_mode: str, min_qty: float = 1.0) -> dict[str, dict[str, Any]]:
    items = build_distributor_compare(result_df, search_mode, min_qty=min_qty)
    out: dict[str, dict[str, Any]] = {}
    for item in items:
        out[str(item["row_key"])] = item
    return out



def build_all_distributor_prices_df(result_df: pd.DataFrame, search_mode: str, min_qty: float = 1.0, price_mode: Optional[str] = None, round100: bool = True, custom_discount: float = 0.0) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if result_df is None or result_df.empty:
        return pd.DataFrame()

    connected_sources = []
    for state_key, label in [("resource_df", "Ресурс"), ("ocs_df", "OCS"), ("merlion_df", "Мерлион")]:
        df = st.session_state.get(state_key)
        if df is not None and not df.empty:
            connected_sources.append(label)

    for _, row in result_df.iterrows():
        article = str(row.get("article", "") or "")
        name = str(row.get("name", "") or "")
        brand = str(row.get("brand", "") or "")
        own_price = float(row.get("sale_price", 0) or 0)
        own_qty = float(row.get("free_qty", 0) or 0)
        selected_price = get_selected_price_raw(row, str(price_mode or "-12%"), bool(round100), float(custom_discount)) if price_mode else None

        rows.append({
            "Артикул": article,
            "Название": name,
            "Производитель": brand,
            "Источник": "Мы",
            "Наша цена": own_price,
            "Наша цена выбранная": selected_price,
            "Наш остаток": own_qty,
            "Цена": own_price,
            "Остаток": own_qty,
            "Разница к нам, руб": 0.0,
            "Разница к нам, %": 0.0,
            "Статус": "наша позиция",
            "Артикул источника": article,
            "Название источника": name,
        })

        choice = {
            "article": article,
            "name": name,
            "brand": brand,
            "sale_price": own_price,
            "row_key": str(row.get("article_norm") or normalize_article(article)),
        }
        offers = {str(offer.get("distributor", "")): offer for offer in get_distributor_offers_for_choice(choice, article, search_mode, min_qty=min_qty)}

        for source_name in connected_sources:
            offer = offers.get(source_name)
            if offer:
                rows.append({
                    "Артикул": article,
                    "Название": name,
                    "Производитель": brand,
                    "Источник": source_name,
                    "Наша цена": own_price,
                    "Наша цена выбранная": selected_price,
                    "Наш остаток": own_qty,
                    "Цена": float(offer.get("price", 0) or 0),
                    "Остаток": float(offer.get("free_qty", 0) or 0),
                    "Разница к нам, руб": float(offer.get("delta", 0) or 0),
                    "Разница к нам, %": round(float(offer.get("delta_percent", 0) or 0), 2),
                    "Статус": str(offer.get("status", "найдено")),
                    "Артикул источника": str(offer.get("article", "") or ""),
                    "Название источника": str(offer.get("name", "") or ""),
                })
            else:
                rows.append({
                    "Артикул": article,
                    "Название": name,
                    "Производитель": brand,
                    "Источник": source_name,
                    "Наша цена": own_price,
                    "Наша цена выбранная": selected_price,
                    "Наш остаток": own_qty,
                    "Цена": pd.NA,
                    "Остаток": pd.NA,
                    "Разница к нам, руб": pd.NA,
                    "Разница к нам, %": pd.NA,
                    "Статус": "нет нормального совпадения",
                    "Артикул источника": "",
                    "Название источника": "",
                })

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    out["_is_own"] = out["Источник"].map(lambda x: 0 if str(x) == "Мы" else 1)
    out = out.sort_values(["Артикул", "_is_own", "Цена", "Источник"], ascending=[True, True, True, True], na_position="last").drop(columns=["_is_own"]).reset_index(drop=True)
    return out



def all_prices_to_excel_bytes(df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Все цены")
    bio.seek(0)
    return bio.read()



def status_visual_class(status: str) -> str:
    status_text = contains_text(status)
    if "ЛУЧШЕ" in status_text:
        return "offer-good"
    if "ДОРОЖЕ" in status_text:
        return "offer-bad"
    if "РАВНА" in status_text:
        return "offer-neutral"
    if "НАША ПОЗИЦИЯ" in status_text:
        return "offer-own"
    return "offer-muted"


def render_results_insight_dashboard(result_df: pd.DataFrame, compare_map: dict[str, dict[str, Any]]) -> None:
    found_count = len(result_df) if isinstance(result_df, pd.DataFrame) else 0
    better_rows = 0
    avg_gain = 0.0
    gains: list[float] = []
    connected = []
    for key, label in [("resource_df", "Ресурс"), ("ocs_df", "OCS"), ("merlion_df", "Мерлион")]:
        df = st.session_state.get(key)
        if isinstance(df, pd.DataFrame) and not df.empty:
            connected.append(label)
    for item in (compare_map or {}).values():
        offer = item.get("best_offer") if isinstance(item, dict) else None
        if offer:
            better_rows += 1
            try:
                gains.append(float(offer.get("delta_percent", 0) or 0))
            except Exception:
                pass
    if gains:
        avg_gain = sum(gains) / len(gains)

    cards = [
        ("🔎", "Найдено позиций", str(found_count), "Сколько строк вошло в текущий результат поиска"),
        ("💚", "Есть цена лучше", str(better_rows), "Сколько позиций можно потенциально пересмотреть по цене"),
        ("📈", "Средняя выгода", (f"{avg_gain:.1f}%" if gains else "—"), "Средняя выгода по тем позициям, где поставщик реально дешевле нас"),
        ("🧩", "Подключено источников", str(len(connected)), (", ".join(connected) if connected else "Файлы дистрибьютеров пока не загружены")),
    ]
    cards_html = "".join(
        f"<div class='insight-card'><div class='insight-top'><span class='insight-icon'>{icon}</span><span class='insight-label'>{label}</span></div><div class='insight-value'>{value}</div><div class='insight-note'>{note}</div></div>"
        for icon, label, value, note in cards
    )
    st.markdown(f"<div class='insight-grid'>{cards_html}</div>", unsafe_allow_html=True)





def build_product_analysis_df(result_df: pd.DataFrame, search_mode: str, min_qty: float = 1.0) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if result_df is None or result_df.empty:
        return pd.DataFrame()

    seen: set[str] = set()
    for _, row in result_df.iterrows():
        row_key = str(row.get("article_norm") or normalize_article(row.get("article", "")))
        if row_key in seen:
            continue
        seen.add(row_key)

        article = str(row.get("article", "") or "")
        name = str(row.get("name", "") or "")
        brand = str(row.get("brand", "") or "")
        own_qty = float(row.get("free_qty", 0) or 0)
        own_price = float(row.get("sale_price", 0) or 0)
        choice = {
            "article": article,
            "name": name,
            "brand": brand,
            "sale_price": own_price,
            "row_key": row_key,
        }
        offers = get_distributor_offers_for_choice(choice, article, search_mode, min_qty=min_qty)
        best_offer = offers[0] if offers else None

        rows.append({
            "Артикул": article,
            "Название": name,
            "Бренд": brand,
            "КОЛ.": own_qty,
            "тек прод": own_price,
            "дистр": float(best_offer.get("price", 0) or 0) if best_offer else None,
            "Дистрибьютор": str(best_offer.get("distributor", "") or "") if best_offer else "",
            "Остаток дистрибьютора": float(best_offer.get("free_qty", 0) or 0) if best_offer else None,
            "Артикул источника": str(best_offer.get("article", "") or "") if best_offer else "",
            "Название источника": str(best_offer.get("name", "") or "") if best_offer else "",
        })

    return pd.DataFrame(rows)



def build_product_analysis_workbook_bytes(result_df: pd.DataFrame, search_mode: str, min_qty: float = 1.0) -> bytes:
    analysis_df = build_product_analysis_df(result_df, search_mode, min_qty=min_qty)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Анализ товара"

    headers = [
        "Артикул", "", "КОЛ.", "тек прод", "дистр", "МИ", "ВЦМ", "Ятовары", "Мы на авито",
        "авито мин", "сред. Зак.", "Прод пред", "пред на Авито", "", "% прод", "% Авито"
    ]
    ws.append(headers)

    column_widths = {
        "A": 14, "B": 4, "C": 10, "D": 12, "E": 12, "F": 10, "G": 10, "H": 12,
        "I": 13, "J": 12, "K": 12, "L": 12, "M": 14, "N": 4, "O": 10, "P": 10,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    header_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9E2F3")
    thin_gray = openpyxl.styles.Side(style="thin", color="D0D7E2")
    border = openpyxl.styles.Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_font = openpyxl.styles.Font(bold=True)
    center = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    currency_format = '#,##0.00'
    percent_format = '0.00%'

    for excel_row, rec in enumerate(analysis_df.to_dict(orient="records"), start=2):
        ws.cell(excel_row, 1).value = rec.get("Артикул", "")
        ws.cell(excel_row, 3).value = rec.get("КОЛ.", None)
        ws.cell(excel_row, 4).value = rec.get("тек прод", None)
        ws.cell(excel_row, 5).value = rec.get("дистр", None)
        ws.cell(excel_row, 6).value = None
        ws.cell(excel_row, 7).value = None
        ws.cell(excel_row, 8).value = None
        ws.cell(excel_row, 9).value = None
        ws.cell(excel_row, 10).value = None
        ws.cell(excel_row, 11).value = None
        ws.cell(excel_row, 12).value = f'=IF(E{excel_row}="","",E{excel_row}-E{excel_row}*5%)'
        ws.cell(excel_row, 13).value = f'=IF(L{excel_row}="","",L{excel_row}-L{excel_row}*20%)'
        ws.cell(excel_row, 15).value = f'=IF(OR(K{excel_row}="",K{excel_row}=0,L{excel_row}=""),"",L{excel_row}/K{excel_row}-1)'
        ws.cell(excel_row, 16).value = f'=IF(OR(K{excel_row}="",K{excel_row}=0,M{excel_row}=""),"",M{excel_row}/K{excel_row}-1)'

        # Comments with context from the parser so the manager sees where the dist price came from.
        if rec.get("дистр") not in (None, ""):
            comment_lines = []
            dist_name = normalize_text(rec.get("Дистрибьютор", ""))
            if dist_name:
                comment_lines.append(f"Лучшее предложение: {dist_name}")
            dist_qty = rec.get("Остаток дистрибьютора")
            if dist_qty not in (None, ""):
                comment_lines.append(f"Остаток: {fmt_qty(dist_qty)} шт.")
            src_article = normalize_text(rec.get("Артикул источника", ""))
            if src_article:
                comment_lines.append(f"Артикул источника: {src_article}")
            src_name = normalize_text(rec.get("Название источника", ""))
            if src_name:
                comment_lines.append(src_name)
            if comment_lines:
                ws.cell(excel_row, 5).comment = openpyxl.comments.Comment("\n".join(comment_lines), "ChatGPT")

        for col_idx in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
            ws.cell(excel_row, col_idx).number_format = currency_format
        for col_idx in [15, 16]:
            ws.cell(excel_row, col_idx).number_format = percent_format

    max_row = max(ws.max_row, 2)
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=16):
        for cell in row:
            cell.border = border
            if cell.column in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16):
                cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{max_row}"

    info = wb.create_sheet("Справка")
    info["A1"] = "Как читать файл"
    info["A1"].font = openpyxl.styles.Font(bold=True, size=12)
    info["A3"] = "Артикул / КОЛ. / тек прод"
    info["B3"] = "Заполняются автоматически из результата поиска и вашего прайса."
    info["A4"] = "дистр"
    info["B4"] = "Подставляется лучшая цена из валидных предложений Ресурс / OCS / Мерлион. В комментарии к ячейке есть дистрибьютор, остаток и источник."
    info["A5"] = "МИ / ВЦМ / Ятовары / Мы на авито / авито мин / сред. Зак."
    info["B5"] = "Эти поля вы заполняете вручную перед обсуждением с руководителем."
    info["A6"] = "Прод пред"
    info["B6"] = "Считается как дистр - 5%."
    info["A7"] = "пред на Авито"
    info["B7"] = "Считается как Прод пред - 20%."
    info["A8"] = "% прод / % Авито"
    info["B8"] = "Считаются относительно среднего закупа."
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 90
    info.freeze_panes = "A3"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()



def render_all_distributor_prices_block(result_df: pd.DataFrame, search_mode: str, min_qty: float, price_mode: str, round100: bool, custom_discount: float) -> None:
    all_prices_df = build_all_distributor_prices_df(
        result_df,
        search_mode,
        min_qty=min_qty,
        price_mode=price_mode,
        round100=round100,
        custom_discount=custom_discount,
    )
    if all_prices_df.empty:
        st.info("Для текущего запроса нет данных по всем ценам дистрибьютеров.")
        return

    st.caption("Здесь видно не только лучшую цену, но и следующую цену у других дистрибьютеров, плюс остаток. Это помогает не снижать цену из-за единичного хвоста на складе.")

    source_order = {"Мы": 0, "Ресурс": 1, "OCS": 2, "Мерлион": 3}
    status_label_map = {
        "offer-good": "🟢 выгоднее",
        "offer-bad": "🔴 дороже",
        "offer-neutral": "🟡 цена равна",
        "offer-own": "🔵 наша позиция",
        "offer-muted": "⚪ без статуса",
    }

    for article, group_df in all_prices_df.groupby("Артикул", sort=False):
        base_name = normalize_text(group_df.iloc[0].get("Название", ""))
        own_row = group_df[group_df["Источник"] == "Мы"].head(1)

        own_price_line = ""
        if not own_row.empty:
            own_price = own_row.iloc[0].get("Цена")
            own_qty = own_row.iloc[0].get("Остаток")
            own_price_line = f"Наша цена: {fmt_price(own_price)} руб. • Остаток: {fmt_qty(own_qty)}"

        st.markdown(
            f"""
            <div class='all-prices-head'>
              <div>
                <div class='all-prices-article'>{html.escape(article)}</div>
                <div class='all-prices-name'>{html.escape(base_name)}</div>
                <div class='all-prices-own'>{own_price_line}</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        work_df = group_df.copy()
        work_df["_rank"] = work_df["Источник"].map(lambda x: source_order.get(str(x), 99))
        work_df = work_df.sort_values(["_rank", "Цена"], na_position="last").reset_index(drop=True)

        cols = st.columns(4)
        for idx, (_, rec) in enumerate(work_df.iterrows()):
            with cols[idx % 4]:
                source = str(rec.get("Источник", "") or "")
                status = str(rec.get("Статус", "") or "")
                status_class = status_visual_class(status)
                badge_text = status_label_map.get(status_class, status or "найдено")

                price_val = rec.get("Цена")
                qty_val = rec.get("Остаток")
                diff_rub = rec.get("Разница к нам, руб")
                diff_pct = rec.get("Разница к нам, %")
                source_article = normalize_text(rec.get("Артикул источника", ""))
                source_name = normalize_text(rec.get("Название источника", ""))

                card_lines = [
                    f"<div class='offer-card-source'>{html.escape(source)}</div>",
                    f"<span class='offer-status-badge {status_class}'>{html.escape(badge_text)}</span>",
                    f"<div class='offer-card-price'>{html.escape(fmt_price(price_val) if pd.notna(price_val) else '—')} {'руб.' if pd.notna(price_val) else ''}</div>",
                    f"<div class='offer-card-meta'>Остаток: <b>{html.escape(fmt_qty(qty_val) if pd.notna(qty_val) else '—')}</b></div>",
                ]

                if source != "Мы" and pd.notna(diff_rub):
                    diff_pct_txt = f" • {round(float(diff_pct), 2):g}%" if pd.notna(diff_pct) else ""
                    sign = "+" if float(diff_rub) < 0 else ""
                    card_lines.append(
                        f"<div class='offer-card-meta'>Разница к нам: {sign}{html.escape(fmt_price(diff_rub))} руб.{html.escape(diff_pct_txt)}</div>"
                    )

                if source_article:
                    card_lines.append(f"<div class='offer-card-code'>{html.escape(source_article)}</div>")
                if source_name:
                    card_lines.append(f"<div class='offer-card-name'>{html.escape(source_name)}</div>")

                st.markdown(
                    "<div class='offer-card-simple'>" + "".join(card_lines) + "</div>",
                    unsafe_allow_html=True,
                )

        show_df = work_df[
            ["Источник", "Цена", "Остаток", "Разница к нам, руб", "Разница к нам, %", "Статус", "Артикул источника", "Название источника"]
        ].copy()
        show_df["Цена"] = show_df["Цена"].apply(lambda v: fmt_price(v) if pd.notna(v) else "")
        show_df["Остаток"] = show_df["Остаток"].apply(lambda v: fmt_qty(v) if pd.notna(v) else "")
        show_df["Разница к нам, руб"] = show_df["Разница к нам, руб"].apply(lambda v: fmt_price(v) if pd.notna(v) else "")
        show_df["Разница к нам, %"] = show_df["Разница к нам, %"].apply(lambda v: (str(round(float(v), 2)).replace(".0", "") + "%") if pd.notna(v) else "")
        with st.expander(f"Таблица по {article}"):
            st.dataframe(show_df, use_container_width=True, hide_index=True, height=min(260, 70 + len(show_df) * 36))

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    st.download_button(
        "⬇️ Скачать все цены в Excel",
        all_prices_to_excel_bytes(all_prices_df),
        file_name="moy_tovar_all_distributor_prices.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def build_full_distributor_report(df: pd.DataFrame, threshold_percent: float, search_mode: str, min_qty: float = 1.0) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    threshold_percent = max(0.0, min(95.0, float(threshold_percent)))
    min_qty = max(0.0, float(min_qty))
    if df is None or df.empty:
        return pd.DataFrame()

    for row in df.itertuples(index=False):
        sale_price = float(getattr(row, "sale_price", 0) or 0)
        own_free_qty = float(getattr(row, "free_qty", 0) or 0)
        if sale_price <= 0 or own_free_qty <= 0:
            continue
        article = str(getattr(row, "article", "") or "")
        name = str(getattr(row, "name", "") or "")
        brand = str(getattr(row, "brand", "") or "")
        choice = {
            "article": article,
            "name": name,
            "brand": brand,
            "sale_price": sale_price,
        }
        best_offer = find_best_distributor_offer_for_choice(choice, article, search_mode, min_qty=min_qty)
        if not best_offer:
            continue
        delta = float(best_offer["delta"])
        delta_percent = ((sale_price - float(best_offer["price"])) / sale_price) * 100.0
        if delta_percent + 1e-9 < threshold_percent:
            continue
        rows.append({
            "Артикул": article,
            "Название": name,
            "Производитель": brand,
            "Наш остаток": float(getattr(row, "free_qty", 0) or 0),
            "Наша цена": sale_price,
            "Лучший дистрибьютер": str(best_offer.get("distributor", "")),
            "Цена дистрибьютора": float(best_offer["price"]),
            "Остаток дистрибьютора": float(best_offer.get("free_qty", 0) or 0),
            "Разница, руб": delta,
            "Разница, %": round(delta_percent, 2),
            "Артикул дистрибьютора": str(best_offer.get("article", "")),
            "Название дистрибьютора": str(best_offer.get("name", "")),
        })
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out = out.sort_values(["Разница, %", "Разница, руб", "Артикул"], ascending=[False, False, True]).reset_index(drop=True)
    return out


def report_to_excel_bytes(df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Отчёт")
    bio.seek(0)
    return bio.read()


def get_series_candidates(df: pd.DataFrame, raw_query: str, series_mode: str = "Только оригиналы") -> dict[str, object]:
    tokens = split_query_parts(raw_query)
    if len(tokens) != 1:
        return {"prefix": "", "candidates": []}
    token = tokens[0]
    token_norm = normalize_article(token)
    if len(token_norm) < 4:
        return {"prefix": token, "candidates": []}
    candidates_by_key: dict[str, dict[str, object]] = {}
    direct_df = df[df["article_norm"].str.startswith(token_norm, na=False)].copy()
    for _, row in direct_df.iterrows():
        candidate = {
            "article": str(row.get("article", "")),
            "article_norm": str(row.get("article_norm", "")),
            "name": str(row.get("name", "")),
            "brand": str(row.get("brand", "")),
            "free_qty": float(row.get("free_qty", 0) or 0),
            "sale_price": float(row.get("sale_price", 0) or 0),
            "is_original": not row_has_negative_series_markers(row),
        }
        candidates_by_key[candidate["article_norm"]] = candidate
    linked_mask = df["name_code_list"].apply(lambda codes: any(str(code).startswith(token_norm) for code in codes))
    linked_df = df[linked_mask].copy()
    for _, row in linked_df.iterrows():
        candidate = {
            "article": str(row.get("article", "")),
            "article_norm": str(row.get("article_norm", "")),
            "name": str(row.get("name", "")),
            "brand": str(row.get("brand", "")),
            "free_qty": float(row.get("free_qty", 0) or 0),
            "sale_price": float(row.get("sale_price", 0) or 0),
            "is_original": not row_has_negative_series_markers(row),
        }
        if candidate["article_norm"] not in candidates_by_key:
            candidates_by_key[candidate["article_norm"]] = candidate
    candidates = list(candidates_by_key.values())
    if series_mode != "Показывать всё":
        original_candidates = [c for c in candidates if bool(c.get("is_original", True))]
        if original_candidates:
            candidates = original_candidates
    candidates.sort(key=series_sort_key)
    if len(candidates) < 2:
        return {"prefix": token, "candidates": []}
    return {"prefix": token, "candidates": candidates}


def build_display_df(df: pd.DataFrame, price_mode: str, round100: bool, custom_discount: float, search_mode: Optional[str] = None, min_qty: float = 1.0) -> pd.DataFrame:
    out = df.copy()
    out["selected_price"] = out.apply(lambda row: get_selected_price_raw(row, price_mode, round100, custom_discount), axis=1)
    label = current_price_label(price_mode, custom_discount)
    display_df = pd.DataFrame(
        {
            "Артикул": out["article"],
            "Название": out["name"],
            "Производитель": out["brand"],
            "Свободно": out["free_qty"].map(fmt_qty),
            "Всего": out["total_qty"].map(fmt_qty),
            "Цена продажи": out["sale_price"].map(fmt_price),
            label: out["selected_price"].map(fmt_price),
            "Алиасы из справочника": out.get("reference_code_list", pd.Series([[] for _ in range(len(out))])).map(lambda x: ", ".join(x) if isinstance(x, list) and x else ""),
        }
    )
    if search_mode and distributor_sources_ready():
        compare_map = distributor_compare_map(df, search_mode, min_qty=min_qty)
        best_dist = []
        best_price = []
        best_qty = []
        best_delta = []
        best_delta_pct = []
        resource_dbg = []
        ocs_dbg = []
        merlion_dbg = []

        def offer_debug_text(offer: dict[str, Any] | None) -> str:
            if not offer:
                return "не найдено/отфильтровано"
            status = str(offer.get("status", "")).strip() or "найдено"
            price = fmt_price(offer.get("price", 0))
            qty = fmt_qty(offer.get("free_qty", 0))
            article = normalize_text(offer.get("article", ""))
            tail = f" • {article}" if article else ""
            return f"{status} • {price} руб. • ост. {qty}{tail}"

        for _, row in df.iterrows():
            item = compare_map.get(str(row.get("article_norm", "")), {})
            offer = item.get("best_offer") if isinstance(item, dict) else None
            if offer:
                best_dist.append(str(offer.get("distributor", "")))
                best_price.append(fmt_price(offer.get("price", 0)))
                best_qty.append(fmt_qty(offer.get("free_qty", 0)))
                best_delta.append(fmt_price(offer.get("delta", 0)))
                best_delta_pct.append(str(offer.get("delta_percent_fmt", "")))
            else:
                best_dist.append("")
                best_price.append("")
                best_qty.append("")
                best_delta.append("")
                best_delta_pct.append("")
            choice = row.to_dict()
            token = str(row.get("article", "") or "")
            resource_dbg.append(offer_debug_text(get_best_distributor_match_for_source(st.session_state.get("resource_df"), choice, token, search_mode, min_qty=min_qty)))
            ocs_dbg.append(offer_debug_text(get_best_distributor_match_for_source(st.session_state.get("ocs_df"), choice, token, search_mode, min_qty=min_qty)))
            merlion_dbg.append(offer_debug_text(get_best_distributor_match_for_source(st.session_state.get("merlion_df"), choice, token, search_mode, min_qty=min_qty)))

        display_df["Лучший дистрибьютер"] = best_dist
        display_df["Цена дистрибьютора"] = best_price
        display_df["Остаток дистрибьютора"] = best_qty
        display_df["Лучше на, руб"] = best_delta
        display_df["Лучше на, %"] = best_delta_pct
        display_df["Ресурс debug"] = resource_dbg
        display_df["OCS debug"] = ocs_dbg
        display_df["Мерлион debug"] = merlion_dbg
    return display_df


def render_avito_open_button(url: str, label: str = "Открыть объявление") -> None:
    if not normalize_text(url):
        st.caption("Ссылка не найдена")
        return
    try:
        st.link_button(label, url, use_container_width=True)
    except Exception:
        st.markdown(f'<a href="{html.escape(url, quote=True)}" target="_blank">{html.escape(label)}</a>', unsafe_allow_html=True)


def render_copy_big_button(text_value: str, button_label: str = "📋 Скопировать весь шаблон") -> None:
    escaped = json.dumps(text_value, ensure_ascii=False)
    html_block = f"""
    <div style='margin-top:8px;'>
      <button onclick='navigator.clipboard.writeText({escaped}).then(() => {{ this.innerText = "Скопировано"; setTimeout(() => this.innerText = {json.dumps(button_label, ensure_ascii=False)}, 1200); }})'
        style='border:none;background:#315efb;color:white;font-weight:800;border-radius:12px;padding:12px 16px;cursor:pointer;min-width:220px;'>
        {html.escape(button_label)}
      </button>
    </div>
    """
    components.html(html_block, height=58)




def render_block_header(title: str, subtitle: str = "", icon: str = "📦", help_text: str = "") -> None:
    tooltip_html = ""
    if normalize_text(help_text):
        tooltip_html = (
            '<div class="block-help-wrap">'
            '<div class="block-help">?</div>'
            f'<div class="block-tooltip">{html.escape(help_text)}</div>'
            '</div>'
        )
    st.markdown(
        f"""
        <div class="block-header">
          <div class="block-header-main">
            <div class="block-icon">{html.escape(icon)}</div>
            <div class="block-title-wrap">
              <div class="block-kicker">Раздел интерфейса</div>
              <div class="section-title">{html.escape(title)}</div>
              <div class="section-sub">{html.escape(subtitle)}</div>
            </div>
          </div>
          <div class="block-header-right">
            <div class="block-sparkles" aria-hidden="true">✦ ✦ ✦</div>
            {tooltip_html}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar_card_header(title: str, icon: str = "📁", help_text: str = "") -> None:
    tooltip_html = ""
    if normalize_text(help_text):
        tooltip_html = (
            '<div class="sidebar-card-help-wrap">'
            '<div class="sidebar-card-help">?</div>'
            f'<div class="sidebar-card-tooltip">{html.escape(help_text)}</div>'
            '</div>'
        )
    st.markdown(
        f"""
        <div class="sidebar-card-header">
          <div class="sidebar-card-header-main">
            <div class="sidebar-card-icon">{html.escape(icon)}</div>
            <div class="sidebar-card-title-wrap">
              <div class="sidebar-card-kicker">Быстрый доступ</div>
              <div class="sidebar-card-title">{html.escape(title)}</div>
            </div>
          </div>
          {tooltip_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_info_banner(title: str, text: str, icon: str = "💡", chips: Optional[list[str]] = None, tone: str = "blue") -> None:
    chips_html = ""
    if chips:
        chips_html = "<div class='banner-chip-row'>" + "".join(
            f"<span class='banner-chip'>{html.escape(chip)}</span>" for chip in chips if normalize_text(chip)
        ) + "</div>"
    st.markdown(
        f"""
        <div class="info-banner tone-{html.escape(tone)}">
          <div class="info-banner-icon">{html.escape(icon)}</div>
          <div class="info-banner-body">
            <div class="info-banner-title">{html.escape(title)}</div>
            <div class="info-banner-text">{html.escape(text)}</div>
            {chips_html}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_action_callout(title: str, text: str, icon: str = "🧭", badges: Optional[list[str]] = None) -> None:
    badges_html = ""
    if badges:
        badges_html = "<div class='callout-badges'>" + "".join(
            f"<span class='callout-badge'>{html.escape(badge)}</span>" for badge in badges if normalize_text(badge)
        ) + "</div>"
    st.markdown(
        f"""
        <div class="action-callout">
          <div class="action-callout-icon">{html.escape(icon)}</div>
          <div class="action-callout-body">
            <div class="action-callout-title">{html.escape(title)}</div>
            <div class="action-callout-text">{html.escape(text)}</div>
            {badges_html}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_report_summary_cards(report_df: pd.DataFrame, threshold_val: float, min_qty_val: float) -> None:
    if report_df is None or report_df.empty:
        return
    max_diff = float(report_df["Разница, %"].max()) if "Разница, %" in report_df.columns else 0.0
    avg_diff = float(report_df["Разница, %"].mean()) if "Разница, %" in report_df.columns else 0.0
    unique_dists = sorted({normalize_text(x) for x in report_df.get("Лучший дистрибьютер", pd.Series(dtype=object)).tolist() if normalize_text(x)})
    cards = [
        ("📦", "Строк в отчёте", str(len(report_df)), "Сколько позиций прошли фильтр по выгоде и остатку"),
        ("🎯", "Порог отбора", f"{fmt_qty(threshold_val)}%", "Минимальная выгода, которую вы задали для отчёта"),
        ("🏷️", "Мин. остаток", f"{fmt_qty(min_qty_val)} шт.", "Отсекает единичные хвосты у дистрибьютеров"),
        ("🚀", "Макс. выгода", f"{max_diff:.1f}%", "Самая сильная разница к вашей цене в текущем отчёте"),
        ("📊", "Средняя выгода", f"{avg_diff:.1f}%", "Средний размер потенциального пересмотра по отчёту"),
        ("🔗", "Чаще лучшие", ", ".join(unique_dists[:3]) if unique_dists else "—", "Какие дистрибьютеры чаще всего дают лучшую цену в текущей выборке"),
    ]
    cards_html = "".join(
        f"<div class='summary-card'><div class='summary-card-top'><span class='summary-card-icon'>{icon}</span><span class='summary-card-label'>{label}</span></div><div class='summary-card-value'>{value}</div><div class='summary-card-note'>{note}</div></div>"
        for icon, label, value, note in cards
    )
    st.markdown(f"<div class='summary-grid'>{cards_html}</div>", unsafe_allow_html=True)

def render_results_table(df: pd.DataFrame, price_mode: str, round100: bool, custom_discount: float, distributor_map: Optional[dict[str, dict[str, Any]]] = None) -> None:
    selected_label = current_price_label(price_mode, custom_discount)
    rows_html = []
    distributor_map = distributor_map or {}
    for _, row in df.iterrows():
        selected_raw = get_selected_price_raw(row, price_mode, round100, custom_discount)
        selected_fmt = fmt_price(selected_raw)
        match_type = str(row.get("match_type", ""))
        row_key = str(row.get("article_norm", ""))
        compare_item = distributor_map.get(row_key, {})
        best_offer = compare_item.get("best_offer") if isinstance(compare_item, dict) else None

        if match_type == "exact":
            badge_html = "<div class='match-badge match-badge-exact'>Точное совпадение</div>"
        elif match_type == "linked":
            badge_html = "<div class='match-badge match-badge-linked'>Найдено по названию</div>"
        else:
            badge_html = "<div class='match-badge match-badge-similar'>Похожее совпадение</div>"

        if best_offer:
            qty_class = "qty-low" if float(best_offer.get("free_qty", 0) or 0) <= 1 else "qty-ok"
            compare_html = f"""
            <div class='best-box'>
              <div class='best-top'>
                <span class='dist-pill'>{html.escape(str(best_offer.get('distributor', '')))}</span>
                <span class='delta-pill'>-{html.escape(str(best_offer.get('delta_percent_fmt', '')))}%</span>
              </div>
              <div class='best-price'>{html.escape(str(best_offer.get('price_fmt', '')))} руб.</div>
              <div class='best-meta'>
                <span class='{qty_class}'>Остаток: {html.escape(fmt_qty(best_offer.get('free_qty', 0)))} шт.</span>
              </div>
              <div class='best-delta'>Лучше на {html.escape(str(best_offer.get('delta_fmt', '')))} руб.</div>
            </div>
            """
        else:
            compare_html = "<div class='best-box best-box-empty'>Нет цены лучше</div>"

        rows_html.append(
            f"""
            <tr>
              <td><span class='article-pill'>{html.escape(str(row['article']))}</span></td>
              <td><div class='name-cell'>{html.escape(str(row['name']))}</div>{badge_html}</td>
              <td>{html.escape(str(row['brand'] or ''))}</td>
              <td>{fmt_qty(row['free_qty'])}</td>
              <td>{fmt_qty(row['total_qty'])}</td>
              <td class='sale-col'>{fmt_price(row['sale_price'])} руб.</td>
              <td class='selected-col'>{selected_fmt}</td>
              <td class='compare-col'>{compare_html}</td>
              <td><button class='copy-btn' onclick="navigator.clipboard.writeText('{selected_fmt}').then(() => {{ this.innerText = 'Скопировано'; setTimeout(() => this.innerText = 'Копировать цену', 1200); }})">Копировать цену</button></td>
            </tr>
            """
        )
    table_html = f"""
    <!doctype html>
    <html><head><meta charset='utf-8'/>
    <style>
      body {{ margin:0; font-family: Inter, Arial, sans-serif; background: transparent; }}
      .wrap {{ background:linear-gradient(180deg, #ffffff 0%, #fbfdff 100%); border:1px solid #dbe5f1; border-radius:22px; overflow:hidden; box-shadow: 0 10px 26px rgba(15,23,42,.06); }}
      table {{ width:100%; border-collapse:separate; border-spacing:0; font-size:14px; }}
      thead th {{ position: sticky; top: 0; z-index: 2; background:linear-gradient(180deg, #f4f8ff 0%, #eef3fb 100%); color:#334155; text-align:left; padding:15px 14px; font-weight:800; border-bottom:1px solid #d7e1ef; }}
      tbody td {{ padding:14px; border-bottom:1px solid #e5edf6; vertical-align:top; color:#1e293b; background: rgba(255,255,255,.96); }}
      tbody tr:nth-child(even) td {{ background: #fcfdff; }}
      tbody tr:hover td {{ background: #f7faff; }}
      tbody tr:last-child td {{ border-bottom:none; }}
      .article-pill {{ display:inline-block; padding:6px 10px; border-radius:999px; background:#edf2ff; color:#315efb; font-weight:800; }}
      .name-cell {{ font-weight:800; line-height:1.35; color:#1e293b; margin-bottom:6px; }}
      .match-badge {{ display:inline-block; padding:5px 10px; border-radius:999px; font-size:12px; font-weight:800; }}
      .match-badge-exact {{ background:#e8f7ee; color:#15803d; }}
      .match-badge-linked {{ background:#e8f1ff; color:#1d4ed8; }}
      .match-badge-similar {{ background:#fff0df; color:#c26a00; }}
      .sale-col {{ font-weight:800; }}
      .selected-col {{ background: linear-gradient(180deg, #f4f8ff 0%, #eef4ff 100%); border-left:1px solid #c7d7ff; border-right:1px solid #c7d7ff; font-weight:900; color:#315efb; white-space:nowrap; box-shadow: inset 0 1px 0 rgba(255,255,255,.7); }}
      .compare-col {{ min-width:220px; }}
      .best-box {{ background:linear-gradient(180deg, #f8fbff 0%, #f3f8ff 100%); border:1px solid #d9e6ff; border-radius:18px; padding:11px 12px; min-width:190px; box-shadow: inset 0 1px 0 rgba(255,255,255,.8); }}
      .best-box-empty {{ color:#64748b; font-weight:700; text-align:center; background:#f8fafc; border-color:#e2e8f0; }}
      .best-top {{ display:flex; justify-content:space-between; gap:8px; align-items:center; margin-bottom:6px; }}
      .dist-pill {{ display:inline-block; padding:5px 10px; border-radius:999px; background:#e9efff; color:#315efb; font-weight:800; }}
      .delta-pill {{ display:inline-block; padding:5px 10px; border-radius:999px; background:#e8f7ee; color:#15803d; font-weight:900; }}
      .best-price {{ font-size:18px; font-weight:900; color:#0f2f83; line-height:1.2; margin-bottom:5px; }}
      .best-meta {{ font-size:12px; margin-bottom:5px; }}
      .qty-low {{ color:#c2410c; font-weight:800; }}
      .qty-ok {{ color:#0f766e; font-weight:800; }}
      .best-delta {{ font-size:12px; color:#64748b; }}
      .copy-btn {{ border:none; background:linear-gradient(180deg, #edf3ff 0%, #e3ecff 100%); color:#315efb; font-weight:800; border-radius:16px; padding:11px 14px; cursor:pointer; min-width:130px; box-shadow: inset 0 1px 0 rgba(255,255,255,.75); }}
    </style></head><body>
      <div class='wrap'><table>
        <thead><tr><th>Артикул</th><th>Название</th><th>Производитель</th><th>Свободно</th><th>Всего</th><th>Цена продажи</th><th>{html.escape(selected_label)}</th><th>Где лучше нас</th><th>Действие</th></tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
      </table></div>
    </body></html>
    """
    height = min(max(180, 70 + len(df) * 84), 1100)
    components.html(table_html, height=height, scrolling=True)


def to_excel_bytes(df: pd.DataFrame, price_mode: str, round100: bool, custom_discount: float, search_mode: Optional[str] = None, min_qty: float = 1.0) -> bytes:
    export_df = build_display_df(df, price_mode, round100, custom_discount, search_mode=search_mode, min_qty=min_qty)
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Результаты")
    bio.seek(0)
    return bio.read()


# ----------------
# Стили интерфейса
# ----------------
st.markdown(
    """
    <style>
    .stApp { background: #eef3f9; }
    header[data-testid="stHeader"] { background: rgba(0,0,0,0); }
    [data-testid="stDecoration"] { display: none; }
    .block-container { max-width: 1560px; padding-top: 3.4rem; padding-bottom: 1.2rem; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #08122f 0%, #102358 55%, #172a63 100%); border-right: 1px solid rgba(255,255,255,.08); }
    [data-testid="stSidebar"] * { color: #e9efff !important; }
    .sidebar-brand { display:flex; align-items:center; gap:12px; margin: 0.15rem 0 0.95rem 0; padding: 0.15rem 0.1rem 0.35rem 0.1rem; }
    .sidebar-brand-logo { width:44px; height:44px; border-radius:14px; background: linear-gradient(180deg, rgba(255,255,255,.18), rgba(255,255,255,.08)); display:flex; align-items:center; justify-content:center; box-shadow: inset 0 1px 0 rgba(255,255,255,.15); font-size:22px; }
    .sidebar-brand-title { font-size: 1.22rem; font-weight: 900; line-height:1.05; color:#ffffff !important; }
    .sidebar-brand-sub { font-size: .82rem; color: #c7d6ff !important; margin-top: 4px; }
    [data-testid="stSidebar"] .stFileUploader section { background: rgba(255,255,255,0.03) !important; border: 1px dashed rgba(255,255,255,0.22) !important; border-radius: 16px !important; padding: 0.6rem !important; }
    [data-testid="stSidebar"] .stFileUploader button, [data-testid="stSidebar"] .stFileUploader button[kind], [data-testid="stSidebar"] .stFileUploader [data-testid="stFileUploaderDropzone"] button, [data-testid="stSidebar"] .stFileUploader [data-testid="baseButton-secondary"], [data-testid="stSidebar"] .stFileUploader [data-baseweb="button"] { background: linear-gradient(180deg, #3767ff 0%, #2455ef 100%) !important; color: #ffffff !important; -webkit-text-fill-color: #ffffff !important; border: none !important; border-radius: 14px !important; font-weight: 800 !important; opacity: 1 !important; box-shadow: 0 10px 20px rgba(49, 94, 251, 0.30) !important; }
    [data-testid="stSidebar"] .stFileUploader small, [data-testid="stSidebar"] .stFileUploader span, [data-testid="stSidebar"] .stFileUploader label { color: #dbe6ff !important; -webkit-text-fill-color: #dbe6ff !important; opacity: 1 !important; }
    [data-testid="stSidebar"] .stButton > button, [data-testid="stSidebar"] .stDownloadButton > button { width: 100% !important; min-height: 48px !important; background: linear-gradient(180deg, #3767ff 0%, #2455ef 100%) !important; color: #ffffff !important; border: none !important; border-radius: 16px !important; font-weight: 900 !important; font-size: 1rem !important; box-shadow: 0 10px 20px rgba(49, 94, 251, 0.30) !important; }
    [data-testid="stSidebar"] .stButton > button:hover, [data-testid="stSidebar"] .stDownloadButton > button:hover { background: linear-gradient(180deg, #4673ff 0%, #2a5cf2 100%) !important; color: #ffffff !important; }
    [data-testid="stSidebar"] .stButton > button:disabled, [data-testid="stSidebar"] .stDownloadButton > button:disabled { background: #5f6f96 !important; color: #edf2ff !important; opacity: .84 !important; box-shadow: none !important; }
    [data-testid="stSidebar"] .stNumberInput input, [data-testid="stSidebar"] .stTextInput input, [data-testid="stSidebar"] .stTextArea textarea, [data-testid="stSidebar"] [data-baseweb="textarea"] textarea, [data-testid="stSidebar"] [data-baseweb="input"] input, [data-testid="stSidebar"] [data-baseweb="base-input"] input, [data-testid="stSidebar"] [data-baseweb="select"] > div { background: #ffffff !important; color: #0f172a !important; -webkit-text-fill-color: #0f172a !important; caret-color: #0f172a !important; border-radius: 16px !important; border: none !important; box-shadow: inset 0 0 0 1px #dbe4f3 !important; }
    [data-testid="stSidebar"] .stTextArea textarea { line-height: 1.55 !important; }
    [data-testid="stSidebar"] .stTextArea textarea::placeholder, [data-testid="stSidebar"] [data-baseweb="textarea"] textarea::placeholder, [data-testid="stSidebar"] .stNumberInput input::placeholder, [data-testid="stSidebar"] .stTextInput input::placeholder { color: #7b8798 !important; -webkit-text-fill-color: #7b8798 !important; opacity: 1 !important; }
    [data-testid="stSidebar"] .stNumberInput button, [data-testid="stSidebar"] .stNumberInput [data-baseweb="button"], [data-testid="stSidebar"] .stNumberInput button svg, [data-testid="stSidebar"] .stNumberInput [data-baseweb="button"] svg { background: #edf3ff !important; color: #1d4ed8 !important; fill: #1d4ed8 !important; stroke: #1d4ed8 !important; border-color: #d9e4ff !important; opacity: 1 !important; }
    [data-testid="stSidebar"] .stRadio > label, [data-testid="stSidebar"] .stSelectbox > label, [data-testid="stSidebar"] .stCheckbox > label, [data-testid="stSidebar"] .stNumberInput > label, [data-testid="stSidebar"] .stTextArea > label, [data-testid="stSidebar"] .stFileUploader > label { color:#ffffff !important; font-weight: 800 !important; font-size: .92rem !important; }
    [data-testid="stSidebar"] .stCheckbox p, [data-testid="stSidebar"] .stRadio p { color:#eef3ff !important; }
    .topbar { position: relative; background: linear-gradient(110deg, #0f172a 0%, #1742a8 56%, #2d6bff 100%); color: white; padding: 18px 20px; border-radius: 24px; margin-top: 0.55rem; margin-bottom: 14px; box-shadow: 0 18px 38px rgba(15, 23, 42, .22); overflow: hidden; }
    .topbar::before { content: ''; position: absolute; inset: -30% auto auto -5%; width: 280px; height: 280px; border-radius: 999px; background: radial-gradient(circle, rgba(255,255,255,.18) 0%, rgba(255,255,255,0) 70%); pointer-events:none; }
    .topbar::after { content: ''; position: absolute; inset: auto -90px -120px auto; width: 240px; height: 240px; border-radius: 999px; background: radial-gradient(circle, rgba(255,255,255,.14) 0%, rgba(255,255,255,0) 72%); pointer-events:none; }
    .topbar-grid { display:grid; grid-template-columns: 1.6fr 1fr 1fr 1fr; gap: 12px; align-items:center; position:relative; z-index:1; }
    .brand-box { display:flex; gap:14px; align-items:center; }
    .logo { width:58px;height:58px;border-radius:18px;background:rgba(255,255,255,.16); display:flex;align-items:center;justify-content:center;font-size:28px;font-weight:700; box-shadow: inset 0 1px 0 rgba(255,255,255,.25), 0 10px 24px rgba(15,23,42,.16); }
    .brand-title { font-size: 25px; font-weight: 900; line-height: 1; letter-spacing: -.02em; }
    .brand-sub { font-size: 13px; opacity: .92; margin-top: 6px; }
    .stat-box { background: rgba(255,255,255,.12); border: 1px solid rgba(255,255,255,.14); border-radius: 18px; padding: 12px 13px; min-height: 76px; backdrop-filter: blur(3px); }
    .stat-cap { font-size: 12px; opacity: .82; margin-bottom: 6px; }
    .stat-val { font-size: 16px; font-weight: 800; line-height: 1.3; }
    .toolbar, .result-wrap { position: relative; background: linear-gradient(180deg, #ffffff 0%, #fbfdff 100%); border: 1px solid #dbe5f1; border-radius: 22px; padding: 16px 18px 18px 18px; margin-bottom: 14px; box-shadow: 0 10px 26px rgba(15, 23, 42, .06); overflow: hidden; transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease; }
    .toolbar:hover, .result-wrap:hover { transform: translateY(-1px); box-shadow: 0 14px 30px rgba(15, 23, 42, .08); border-color: #cfe0ff; }
    .toolbar::before, .result-wrap::before { content: ''; position: absolute; inset: 0 auto auto 0; width: 100%; height: 4px; background: linear-gradient(90deg, #315efb 0%, #79a6ff 100%); opacity: .95; }
    .toolbar::after, .result-wrap::after { content: ''; position: absolute; right: -65px; top: -65px; width: 180px; height: 180px; border-radius: 999px; background: radial-gradient(circle, rgba(49,94,251,.08) 0%, rgba(49,94,251,0) 72%); pointer-events:none; }
    .block-header { display:flex; align-items:flex-start; justify-content:space-between; gap:16px; padding: 2px 0 14px 0; margin-bottom: 14px; border-bottom: 1px solid #e7eef9; position: relative; }
    .block-header-main { display:flex; align-items:flex-start; gap:14px; min-width: 0; }
    .block-header-right { display:flex; align-items:flex-start; gap:10px; flex: 0 0 auto; }
    .block-icon { width: 48px; height: 48px; border-radius: 16px; background: linear-gradient(180deg, #3767ff 0%, #2455ef 100%); color: #ffffff; display:flex; align-items:center; justify-content:center; font-size: 24px; flex: 0 0 48px; box-shadow: 0 12px 22px rgba(49, 94, 251, .22); position: relative; }
    .block-icon::after { content: ''; position: absolute; inset: auto -4px -5px auto; width: 16px; height: 16px; border-radius: 999px; background: rgba(255,255,255,.25); }
    .block-title-wrap { min-width: 0; }
    .block-kicker { display:inline-flex; align-items:center; padding: 4px 9px; margin-bottom: 7px; border-radius: 999px; background: #eef4ff; border: 1px solid #d8e5ff; color: #315efb; font-size: 11px; font-weight: 900; letter-spacing: .04em; text-transform: uppercase; }
    .toolbar-title, .section-title { font-size: 22px; font-weight: 900; color:#0f172a; margin:0 0 5px 0; line-height:1.12; letter-spacing:-0.02em; }
    .toolbar-sub, .section-sub { font-size: 13px; color:#64748b; margin:0; line-height:1.55; max-width: 980px; }
    .block-sparkles { display:flex; align-items:center; gap: 3px; color:#89a9ff; font-size: 12px; font-weight: 900; letter-spacing: .04em; opacity: .9; margin-top: 5px; }
    .block-help-wrap { position: relative; flex: 0 0 auto; }
    .block-help { display:flex; align-items:center; justify-content:center; width: 32px; height: 32px; border-radius: 999px; border: 1px solid #cfe0ff; background: linear-gradient(180deg, #f6f9ff 0%, #eef4ff 100%); color: #315efb; font-size: 15px; font-weight: 900; cursor: help; user-select: none; box-shadow: inset 0 1px 0 rgba(255,255,255,.8), 0 6px 14px rgba(49,94,251,.08); }
    .block-help:hover { background: #eaf1ff; transform: translateY(-1px); }
    .block-tooltip { position: absolute; right: 0; top: 40px; width: 340px; max-width: min(340px, 82vw); padding: 13px 14px; border-radius: 16px; background: #0f172a; color: #f8fbff; font-size: 12.8px; line-height: 1.5; box-shadow: 0 18px 36px rgba(15, 23, 42, .28); opacity: 0; transform: translateY(6px); pointer-events: none; transition: opacity .18s ease, transform .18s ease; z-index: 20; }
    .block-tooltip::before { content: ''; position: absolute; top: -6px; right: 10px; width: 12px; height: 12px; background: #0f172a; transform: rotate(45deg); }
    .block-help-wrap:hover .block-tooltip { opacity: 1; transform: translateY(0); }
    .sidebar-card { background: linear-gradient(180deg, rgba(255,255,255,.08), rgba(255,255,255,.045)); border: 1px solid rgba(255,255,255,.13); border-radius: 22px; padding: 1rem 0.95rem 0.95rem 0.95rem; margin: 0.95rem 0 1.05rem 0; box-shadow: 0 12px 26px rgba(2, 8, 23, .24), inset 0 1px 0 rgba(255,255,255,.06); position: relative; overflow: hidden; }
    .sidebar-card::before { content: ''; position: absolute; inset: 0 auto auto 0; width: 100%; height: 3px; background: linear-gradient(90deg, rgba(111,163,255,.95) 0%, rgba(49,94,251,.95) 100%); opacity: .95; }
    .sidebar-card-header { display:flex; align-items:flex-start; justify-content:space-between; gap:10px; margin-bottom: .6rem; padding-bottom: .55rem; border-bottom: 1px solid rgba(255,255,255,.10); }
    .sidebar-card-header-main { display:flex; align-items:center; gap:10px; min-width:0; }
    .sidebar-card-title-wrap { min-width: 0; }
    .sidebar-card-kicker { color:#cfe0ff !important; font-size:10px; text-transform: uppercase; letter-spacing:.06em; font-weight:900; margin-bottom:2px; }
    .sidebar-card-icon { width:34px; height:34px; border-radius:12px; background: linear-gradient(180deg, rgba(255,255,255,.18), rgba(255,255,255,.08)); display:flex; align-items:center; justify-content:center; font-size:17px; box-shadow: inset 0 1px 0 rgba(255,255,255,.12); flex: 0 0 34px; }
    .sidebar-card-title { font-size: 1.01rem; font-weight: 900; color:#ffffff !important; line-height:1.15; margin:0; }
    .sidebar-card-help-wrap { position: relative; flex: 0 0 auto; }
    .sidebar-card-help { display:flex; align-items:center; justify-content:center; width:24px; height:24px; border-radius:999px; border:1px solid rgba(255,255,255,.18); background: rgba(255,255,255,.08); color:#ffffff !important; font-size:12px; font-weight:900; cursor:help; user-select:none; }
    .sidebar-card-tooltip { position:absolute; right:0; top:30px; width:250px; max-width:min(250px, 66vw); padding:10px 11px; border-radius:12px; background:#f8fbff; color:#0f172a !important; font-size:12px; line-height:1.45; box-shadow:0 16px 34px rgba(2, 8, 23, .30); opacity:0; transform:translateY(6px); pointer-events:none; transition:opacity .18s ease, transform .18s ease; z-index:35; }
    .sidebar-card-tooltip::before { content:''; position:absolute; top:-6px; right:9px; width:12px; height:12px; background:#f8fbff; transform:rotate(45deg); }
    .sidebar-card-help-wrap:hover .sidebar-card-tooltip { opacity:1; transform:translateY(0); }
    .sidebar-card-note { font-size: .79rem; line-height: 1.52; color:#c7d6ff !important; margin-bottom: .65rem; }
    .sidebar-status { background: rgba(7, 31, 74, .92); border: 1px solid rgba(255,255,255,.06); border-radius: 14px; padding: .76rem .82rem; color:#ffffff !important; font-weight: 800; margin-top: .58rem; }
    .sidebar-mini { font-size:.78rem; color:#c7d6ff !important; line-height:1.5; margin-top:.65rem; }
    .stButton > button, .stDownloadButton > button { min-height: 48px !important; border-radius: 16px !important; font-weight: 900 !important; border: none !important; box-shadow: 0 10px 18px rgba(49,94,251,.12) !important; transition: transform .16s ease, box-shadow .16s ease, filter .16s ease !important; }
    .stButton > button:hover, .stDownloadButton > button:hover { transform: translateY(-1px); box-shadow: 0 14px 24px rgba(49,94,251,.16) !important; filter: saturate(1.02); }
    .stButton > button:focus, .stDownloadButton > button:focus { box-shadow: 0 0 0 3px rgba(49,94,251,.12), 0 12px 22px rgba(49,94,251,.16) !important; }
    div[data-testid="metric-container"] { background: linear-gradient(180deg, #ffffff 0%, #f9fbff 100%); border: 1px solid #dbe7fb; border-radius: 18px; padding: 10px 12px; box-shadow: inset 0 1px 0 rgba(255,255,255,.7); }
    div[data-testid="metric-container"] > label { color:#64748b !important; font-weight:700 !important; }
    div[data-testid="metric-container"] [data-testid="stMetricValue"] { color:#0f172a !important; font-weight:900 !important; }
    [data-testid="stExpander"] { border: 1px solid #dbe5f1; border-radius: 16px; overflow: hidden; background: linear-gradient(180deg, #ffffff 0%, #fbfdff 100%); }
    [data-testid="stExpander"] details summary { background: #f8fbff; border-bottom: 1px solid #e6eefb; }
    [data-testid="stDataFrame"] { border: 1px solid #dbe5f1; border-radius: 18px; overflow: hidden; box-shadow: 0 6px 18px rgba(15,23,42,.05); }
    [data-testid="stForm"] { background: linear-gradient(180deg, rgba(255,255,255,.35), rgba(255,255,255,.1)); border-radius: 18px; }
    .mini-chip { display:inline-flex; align-items:center; gap:6px; padding:7px 11px; border-radius:999px; background:#eef4ff; color:#315efb; font-weight:800; font-size:12px; margin-right:6px; margin-bottom:6px; border: 1px solid #d6e3ff; }
    .soft-note { margin: 8px 0 12px 0; padding: 11px 14px; border-radius: 16px; background: linear-gradient(180deg, #f7fbff 0%, #eef5ff 100%); border: 1px solid #d6e4ff; color: #44607f; font-size: 13px; line-height: 1.55; }
    .result-inline-stat { display:inline-flex; align-items:center; gap:8px; margin: 2px 0 14px 0; padding: 9px 13px; border-radius: 999px; background:#eefaf1; color:#166534; border:1px solid #cbead4; font-weight:800; }
    .insight-grid { display:grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; margin: 14px 0 16px 0; }
    .insight-card { background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%); border: 1px solid #dbe7fb; border-radius: 20px; padding: 14px 15px; box-shadow: 0 8px 18px rgba(15,23,42,.05); }
    .insight-top { display:flex; align-items:center; gap:8px; margin-bottom: 10px; }
    .insight-icon { width:32px; height:32px; display:flex; align-items:center; justify-content:center; border-radius: 12px; background:#eef4ff; font-size:16px; }
    .insight-label { color:#64748b; font-size:12px; font-weight:800; }
    .insight-value { color:#0f172a; font-size: 28px; font-weight: 900; line-height:1.1; margin-bottom: 6px; }
    .insight-note { color:#6b7c93; font-size:12px; line-height:1.45; }
    .all-prices-head { display:flex; align-items:flex-start; justify-content:space-between; gap:10px; margin: 14px 0 10px 0; padding: 14px 16px; border-radius: 18px; background: linear-gradient(180deg, #fbfdff 0%, #f5f9ff 100%); border:1px solid #dbe7fb; }
    .all-prices-article { color:#315efb; font-size: 18px; font-weight: 900; margin-bottom: 4px; }
    .all-prices-name { color:#0f172a; font-size: 14px; font-weight: 800; line-height: 1.45; }
    .all-prices-own { margin-top: 6px; color:#64748b; font-size: 12.5px; }
    .offers-grid { display:grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 12px; margin: 0 0 14px 0; }
    .offer-card { border-radius: 18px; padding: 14px; border:1px solid #dbe7fb; background: linear-gradient(180deg, #ffffff 0%, #f9fbff 100%); box-shadow: 0 8px 18px rgba(15,23,42,.05); min-height: 172px; }
    .offer-card-top { display:flex; align-items:flex-start; justify-content:space-between; gap:8px; margin-bottom: 10px; }
    .offer-source { color:#0f172a; font-size: 15px; font-weight: 900; }
    .offer-status { display:inline-flex; align-items:center; justify-content:center; padding:5px 9px; border-radius:999px; font-size:11px; font-weight:900; }
    .offer-good .offer-status { background:#e9f9ef; color:#15803d; }
    .offer-bad .offer-status { background:#fff1f2; color:#be123c; }
    .offer-neutral .offer-status { background:#eef4ff; color:#315efb; }
    .offer-own .offer-status { background:#f3f4f6; color:#475569; }
    .offer-muted .offer-status { background:#f8fafc; color:#64748b; }
    .offer-price { color:#0f2f83; font-size: 24px; font-weight: 900; line-height: 1.15; margin-bottom: 6px; }
    .offer-meta { color:#64748b; font-size: 12.5px; line-height:1.45; margin-bottom: 4px; }
    .offer-code { color:#315efb; font-size:12px; font-weight:800; margin-top: 8px; }
    .offer-name { color:#334155; font-size:12px; line-height:1.45; margin-top:4px; }
    @media (max-width: 1100px) {
        .insight-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
        .offers-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 700px) {
        .insight-grid, .offers-grid { grid-template-columns: 1fr; }
    }
    .info-banner { display:flex; gap:14px; align-items:flex-start; padding:15px 16px; margin: 6px 0 14px 0; border-radius: 18px; border: 1px solid #dbe7fb; background: linear-gradient(180deg, #fbfdff 0%, #f5f9ff 100%); box-shadow: 0 8px 18px rgba(15,23,42,.05); }
    .info-banner-icon { width:42px; height:42px; flex:0 0 42px; border-radius: 14px; display:flex; align-items:center; justify-content:center; font-size: 20px; background: linear-gradient(180deg, #3767ff 0%, #2455ef 100%); color:#fff; box-shadow: 0 10px 18px rgba(49,94,251,.18); }
    .info-banner-body { min-width: 0; }
    .info-banner-title { font-size: 15px; font-weight: 900; color:#0f172a; margin-bottom: 4px; }
    .info-banner-text { font-size: 13px; line-height: 1.55; color:#64748b; }
    .banner-chip-row { display:flex; flex-wrap:wrap; gap:8px; margin-top: 10px; }
    .banner-chip { display:inline-flex; align-items:center; gap:6px; padding: 6px 10px; border-radius: 999px; background:#eef4ff; border:1px solid #d8e5ff; color:#315efb; font-size: 12px; font-weight: 800; }
    .tone-green { background: linear-gradient(180deg, #fbfffd 0%, #f2fff7 100%); border-color: #d2f1dd; }
    .tone-green .info-banner-icon { background: linear-gradient(180deg, #16a34a 0%, #15803d 100%); box-shadow: 0 10px 18px rgba(22,163,74,.18); }
    .tone-purple { background: linear-gradient(180deg, #fcfbff 0%, #f6f3ff 100%); border-color: #e6dcff; }
    .tone-purple .info-banner-icon { background: linear-gradient(180deg, #7c3aed 0%, #6d28d9 100%); box-shadow: 0 10px 18px rgba(124,58,237,.18); }
    .action-callout { display:flex; gap:14px; align-items:flex-start; padding:16px 17px; margin: 12px 0 10px 0; border-radius: 20px; background: linear-gradient(135deg, #0f172a 0%, #1d4ed8 100%); color:#ffffff; box-shadow: 0 18px 34px rgba(29,78,216,.18); position: relative; overflow:hidden; }
    .action-callout::after { content:'✦'; position:absolute; right:16px; top:8px; font-size:42px; color: rgba(255,255,255,.08); transform: rotate(12deg); }
    .action-callout-icon { width:46px; height:46px; flex:0 0 46px; border-radius: 15px; background: rgba(255,255,255,.16); display:flex; align-items:center; justify-content:center; font-size: 22px; }
    .action-callout-title { font-size: 15px; font-weight: 900; margin-bottom: 4px; }
    .action-callout-text { font-size: 13px; line-height:1.55; color: rgba(255,255,255,.9); }
    .callout-badges { display:flex; flex-wrap:wrap; gap:8px; margin-top: 10px; }
    .callout-badge { display:inline-flex; align-items:center; padding: 6px 10px; border-radius:999px; background: rgba(255,255,255,.12); border:1px solid rgba(255,255,255,.14); font-size:12px; font-weight:800; color:#fff; }
    .summary-grid { display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin: 12px 0 14px 0; }
    .summary-card { background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%); border:1px solid #dbe7fb; border-radius: 18px; padding: 14px 15px; box-shadow: 0 8px 18px rgba(15,23,42,.05); }
    .summary-card-top { display:flex; align-items:center; gap:8px; margin-bottom:10px; }
    .summary-card-icon { width:32px; height:32px; border-radius:12px; display:flex; align-items:center; justify-content:center; background:#eef4ff; font-size:16px; }
    .summary-card-label { color:#64748b; font-size:12px; font-weight:800; }
    .summary-card-value { color:#0f172a; font-size: 24px; font-weight:900; line-height:1.15; margin-bottom: 6px; }
    .summary-card-note { color:#6b7c93; font-size:12px; line-height:1.45; }
    div[data-testid="stExpander"] details { border:1px solid #dbe7fb; border-radius: 18px; background: linear-gradient(180deg, #ffffff 0%, #fbfdff 100%); overflow:hidden; box-shadow: 0 8px 18px rgba(15,23,42,.04); }
    div[data-testid="stExpander"] summary { padding-top: 4px; padding-bottom: 4px; }
    div[data-testid="stDataFrame"] { border-radius: 18px; overflow:hidden; border: 1px solid #e2e8f0; box-shadow: 0 8px 18px rgba(15,23,42,.04); }
    .analysis-help-grid { display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 10px; margin: 10px 0 14px 0; }
    .analysis-help-card { border-radius: 16px; border:1px solid #dbe7fb; background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%); padding: 12px 13px; }
    .analysis-help-title { font-size: 13px; font-weight: 900; color:#0f172a; margin-bottom: 5px; }
    .analysis-help-text { font-size: 12px; line-height:1.5; color:#64748b; }
    @media (max-width: 1100px) {
        .summary-grid, .analysis-help-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 700px) {
        .summary-grid, .analysis-help-grid { grid-template-columns: 1fr; }
        .info-banner, .action-callout { padding: 14px; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# -------
# Sidebar
# -------
with st.sidebar:
    st.markdown(
        """
        <div class="sidebar-brand">
          <div class="sidebar-brand-logo">📦</div>
          <div>
            <div class="sidebar-brand-title">Мой Товар</div>
            <div class="sidebar-brand-sub">Почти как локальная версия 💙</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Загрузить прайс", "📥", "Основной прайс каталога. Из него приложение берёт артикул, название, остаток и цену продажи.")
    uploaded = st.file_uploader("Загрузить прайс", type=["xlsx", "xls", "xlsm", "csv"], label_visibility="collapsed")
    if uploaded is not None:
        try:
            st.session_state.catalog_base_df = load_price_file(uploaded.name, uploaded.getvalue())
            st.session_state.catalog_name = uploaded.name
            rebuild_catalog_effective_df()
        except Exception as exc:
            st.error(f"Ошибка: {exc}")
    file_caption = st.session_state.get("catalog_name", "Файл ещё не выбран")
    st.markdown(f'<div class="sidebar-status">Загружен: {html.escape(file_caption)}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Справочник артикулов", "🧠", "Дополнительный словарь коротких и длинных артикулов. Используется в самом конце, если обычного поиска и кодов из названия недостаточно.")
    st.markdown('<div class="sidebar-card-note">Необязательный файл. Используется в самом конце как словарь соответствий между короткими и длинными артикулами.</div>', unsafe_allow_html=True)
    article_ref_uploaded = st.file_uploader("Загрузить справочник артикулов", type=["xlsx", "xls", "xlsm", "csv"], key="article_ref_uploader", label_visibility="collapsed")
    if article_ref_uploaded is not None:
        try:
            st.session_state.article_ref_df = load_article_reference_file(article_ref_uploaded.name, article_ref_uploaded.getvalue())
            st.session_state.article_ref_name = article_ref_uploaded.name
            rebuild_catalog_effective_df()
            submitted_query = normalize_text(st.session_state.get("submitted_query", ""))
            if submitted_query and isinstance(st.session_state.catalog_df, pd.DataFrame):
                st.session_state.last_result = perform_search(st.session_state.catalog_df, submitted_query, st.session_state.get("search_mode", "Только артикул"))
        except Exception as exc:
            st.error(f"Ошибка справочника: {exc}")
    article_ref_caption = st.session_state.get("article_ref_name", "ещё не загружен")
    st.markdown(f'<div class="sidebar-status">Справочник: {html.escape(article_ref_caption)}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Загрузить файл Авито", "🛒", "Файл с объявлениями Авито. Помогает быстро понять, есть ли у найденной позиции действующее объявление и открыть его.")
    st.markdown('<div class="sidebar-card-note">Файл с колонкой <b>Название объявления</b>. Ссылки можно читать прямо из гиперссылок Excel.</div>', unsafe_allow_html=True)
    avito_uploaded = st.file_uploader("Загрузить файл Авито", type=["xlsx", "xlsm", "csv"], key="avito_uploader", label_visibility="collapsed")
    if avito_uploaded is not None:
        try:
            st.session_state.avito_df = load_avito_file(avito_uploaded.name, avito_uploaded.getvalue())
            st.session_state.avito_name = avito_uploaded.name
        except Exception as exc:
            st.error(f"Ошибка файла Авито: {exc}")
    avito_caption = st.session_state.get("avito_name", "ещё не загружен")
    st.markdown(f'<div class="sidebar-status">Авито: {html.escape(avito_caption)}</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Дистрибьютеры", "🏷️", "Здесь подключаются прайсы Ресурс, OCS и Мерлион. Блок сравнивает только нормальные оригинальные позиции и только то, что есть в наличии.")
    st.markdown('<div class="sidebar-card-note">Добавлен перенос логики сравнения цен: только оригиналы, только хорошие позиции, только товар в наличии.</div>', unsafe_allow_html=True)
    resource_uploaded = st.file_uploader("Ресурс", type=["xlsx", "xlsm"], key="resource_uploader")
    if resource_uploaded is not None:
        try:
            st.session_state.resource_df = load_resource_file(resource_uploaded.name, resource_uploaded.getvalue())
            st.session_state.resource_name = resource_uploaded.name
        except Exception as exc:
            st.error(f"Ошибка файла Ресурс: {exc}")
    st.markdown(f'<div class="sidebar-status">Ресурс: {html.escape(st.session_state.get("resource_name", "ещё не загружен"))}</div>', unsafe_allow_html=True)

    ocs_uploaded = st.file_uploader("OCS", type=["xlsx", "xlsm"], key="ocs_uploader")
    if ocs_uploaded is not None:
        try:
            st.session_state.ocs_df = load_ocs_file(ocs_uploaded.name, ocs_uploaded.getvalue())
            st.session_state.ocs_name = ocs_uploaded.name
        except Exception as exc:
            st.error(f"Ошибка файла OCS: {exc}")
    st.markdown(f'<div class="sidebar-status">OCS: {html.escape(st.session_state.get("ocs_name", "ещё не загружен"))}</div>', unsafe_allow_html=True)

    merlion_uploaded = st.file_uploader("Мерлион", type=["xlsx", "xlsm"], key="merlion_uploader")
    if merlion_uploaded is not None:
        try:
            st.session_state.merlion_df = load_merlion_file(merlion_uploaded.name, merlion_uploaded.getvalue())
            st.session_state.merlion_name = merlion_uploaded.name
        except Exception as exc:
            st.error(f"Ошибка файла Мерлион: {exc}")
    st.markdown(f'<div class="sidebar-status">Мерлион: {html.escape(st.session_state.get("merlion_name", "ещё не загружен"))}</div>', unsafe_allow_html=True)
    st.number_input("Порог отчёта, %", min_value=0.0, max_value=95.0, step=1.0, key="distributor_threshold")
    st.number_input("Мин. остаток у дистрибьютора", min_value=1.0, max_value=999999.0, step=1.0, key="distributor_min_qty")
    st.markdown('<div class="sidebar-mini">Если у поставщика осталась 1 шт., можно поднять минимальный остаток и убрать такие хвосты из сравнения.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Быстрая правка цен", "✏️", "Локально обновляет цены в загруженном прайсе по вставленному списку артикулов и новых цен. Удобно для быстрых правок без изменения исходного файла.")
    st.markdown('<div class="sidebar-card-note">Вставьте строки вида <b>CE278A 8900</b>, <b>CF364A - 29700</b> или прямо текст из Telegram.</div>', unsafe_allow_html=True)
    st.text_area(
        "Вставьте строки вроде: CE278A 8900",
        key="price_patch_input",
        height=110,
        label_visibility="collapsed",
        placeholder="""CE278A 8900
CE278AC 7900
CF364A - 29700 🔽""",
    )
    if st.button("Править цены в прайсе", use_container_width=True):
        if isinstance(st.session_state.catalog_base_df, pd.DataFrame):
            updated_base_df, patch_message = apply_price_updates(st.session_state.catalog_base_df, st.session_state.price_patch_input)
            st.session_state.catalog_base_df = updated_base_df
            rebuild_catalog_effective_df()
            st.session_state.patch_message = patch_message
            submitted_query = normalize_text(st.session_state.get("submitted_query", ""))
            if submitted_query and isinstance(st.session_state.catalog_df, pd.DataFrame):
                st.session_state.last_result = perform_search(st.session_state.catalog_df, submitted_query, st.session_state.get("search_mode", "Только артикул"))
        else:
            st.session_state.patch_message = "Сначала загрузите прайс."
    if st.session_state.patch_message:
        st.markdown(f'<div class="sidebar-mini">{html.escape(st.session_state.patch_message)}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="sidebar-mini">Прайс сохраняется локально. После правок цены не пропадут до загрузки нового файла.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Настройки", "⚙️", "Управляет режимом поиска, основной ценой, пользовательской скидкой и округлением. На ядро поиска не влияет, только на отображение и шаблоны.")
    st.selectbox("Режим поиска", ["Только артикул", "Умный", "Артикул + название + бренд"], key="search_mode")
    st.radio("Какая цена главная", ["-12%", "-20%", "Своя скидка"], key="price_mode")
    st.number_input("Своя скидка, %", min_value=0.0, max_value=99.0, step=1.0, key="custom_discount")
    st.checkbox("Округлять вверх до 100", key="round100")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="sidebar-card">', unsafe_allow_html=True)
    render_sidebar_card_header("Текст шаблона 1", "🧾", "Постоянный хвост для первого шаблона. Здесь можно хранить адрес, условия работы и доставку — они подставляются автоматически в конец шаблона.")
    st.markdown('<div class="sidebar-card-note">Этот текст добавляется один раз в конце шаблона 1. Хэштеги по артикулам подставляются автоматически.</div>', unsafe_allow_html=True)
    st.text_area("Текст шаблона 1", key="template1_footer", height=170, label_visibility="collapsed")
    st.markdown('<div class="sidebar-mini">Текст сохраняется локально и останется до следующего изменения.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


# ---------
# Верхняя панель
# ---------
catalog_df = st.session_state.get("catalog_df")
file_name = st.session_state.get("catalog_name", "ещё не загружен")
rows_count = len(catalog_df) if isinstance(catalog_df, pd.DataFrame) else 0
price_mode = st.session_state.price_mode
round100 = st.session_state.round100
custom_discount = float(st.session_state.custom_discount)
search_mode = st.session_state.search_mode
price_label = current_price_label(price_mode, custom_discount)

st.markdown(f"""
<div class="topbar"><div class="topbar-grid">
<div class="brand-box"><div class="logo">📦</div><div><div class="brand-title">{APP_TITLE}</div><div class="brand-sub">Streamlit • поиск • шаблоны • правка цен • сравнение с дистрибьюторами</div></div></div>
<div class="stat-box"><div class="stat-cap">Текущий прайс</div><div class="stat-val">{html.escape(file_name)}</div></div>
<div class="stat-box"><div class="stat-cap">Строк в каталоге</div><div class="stat-val">{rows_count}</div></div>
<div class="stat-box"><div class="stat-cap">Режим цены</div><div class="stat-val">{html.escape(price_label)}{' • округл.' if round100 else ''}</div></div>
</div></div>
""", unsafe_allow_html=True)


# ------
# Поиск
# ------
st.markdown('<div class="toolbar">', unsafe_allow_html=True)
render_block_header(
    "Поиск товара",
    "Можно искать по одному или нескольким артикулам. Пробелы, /, запятые и Enter тоже поддерживаются.",
    icon="🔎",
    help_text="Основной блок поиска по вашему прайсу. Сначала ищет точное совпадение по артикулу, затем связанные коды из названия и только потом — более мягкие варианты, если они разрешены режимом поиска.",
)

with st.form("search_form", clear_on_submit=False):
    search_value = st.text_area(
        "Поисковый запрос",
        value=st.session_state.search_input,
        placeholder="Например:\nCC530AC CC531AC CC532AC\nили\n842025 / 841913 / 841711 / 842339\nили Xerox 700",
        height=90,
        label_visibility="collapsed",
    )
    c1, c2, c3 = st.columns([1, 1, 2.4])
    find_clicked = c1.form_submit_button("🔎 Найти", use_container_width=True, type="primary")
    clear_clicked = c2.form_submit_button("🧹 Очистить", use_container_width=True)
    c3.markdown("<div style='padding-top:9px;color:#64748b;font-size:12px;'>Если код не найден в колонке «Артикул», приложение пробует найти его как связанный код в названии позиции.</div>", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

if clear_clicked:
    st.session_state.search_input = ""
    st.session_state.submitted_query = ""
    st.session_state.last_result = None
    st.rerun()

if find_clicked:
    normalized_query = normalize_query_for_display(search_value)
    st.session_state.search_input = normalized_query
    st.session_state.submitted_query = normalized_query
    st.session_state.last_result = (perform_search(st.session_state.catalog_df, normalized_query, search_mode) if isinstance(st.session_state.catalog_df, pd.DataFrame) else None)
    st.rerun()

current_df = st.session_state.catalog_df
submitted_query = st.session_state.submitted_query
result_df = st.session_state.last_result
min_dist_qty = float(st.session_state.distributor_min_qty)


# ----------
# Результаты
# ----------
st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
render_block_header(
    "Результаты",
    "Точное совпадение — по колонке «Артикул». Найдено по названию — когда код сидит в названии той же позиции.",
    icon="📊",
    help_text="Здесь показываются найденные позиции из вашего прайса. Если подключены дистрибьютеры, рядом появляется блок «Где лучше нас» с лучшей более дешёвой ценой поставщика, остатком и выгодой.",
)

if current_df is None:
    st.info("Сначала загрузите прайс в левой панели 👈")
elif not submitted_query.strip():
    st.info("Введите артикул или название и нажмите **Найти**.")
elif result_df is None or result_df.empty:
    st.warning("Ничего не найдено. Попробуйте другой артикул, бренд или часть названия.")
else:
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Найдено", len(result_df))
    m2.metric("Цена", price_label)
    m3.metric("Округление", "вкл" if round100 else "выкл")
    m4.metric("Каталог", len(current_df))

    compare_map = distributor_compare_map(result_df, search_mode, min_qty=min_dist_qty) if distributor_sources_ready() else {}
    render_results_insight_dashboard(result_df, compare_map)
    if compare_map:
        render_info_banner(
            "Как читать результаты сравнения",
            "В таблице выше показывается только поставщик, который реально дешевле вас. Если блок пустой, это не всегда означает, что совпадений нет — часто это значит, что поставщики нашлись, но просто дороже вашей цены.",
            icon="🧠",
            chips=["пустой блок ≠ нет совпадения", "остаток уже учтён", "плохие предложения отсеяны"],
            tone="green",
        )
        better_rows = sum(1 for item in compare_map.values() if item.get("best_offer"))
        chips = []
        if st.session_state.get("resource_df") is not None:
            chips.append("<span class=\"mini-chip\">Ресурс подключён</span>")
        if st.session_state.get("ocs_df") is not None:
            chips.append("<span class=\"mini-chip\">OCS подключён</span>")
        if st.session_state.get("merlion_df") is not None:
            chips.append("<span class=\"mini-chip\">Мерлион подключён</span>")
        st.markdown("".join(chips), unsafe_allow_html=True)
        st.caption(f"Для найденных позиций проверяю только оригиналы, только хорошие предложения и только остатки от {fmt_qty(min_dist_qty)} шт. Где цена поставщика лучше — показываю дистрибьютора, цену, остаток и выгоду.")
        st.metric("Где кто-то лучше нас", better_rows)

    render_results_table(result_df.head(200), price_mode, round100, custom_discount, distributor_map=compare_map)
    with st.expander("Показать техническую таблицу"):
        render_info_banner(
            "Техническая таблица для проверки логики",
            "Этот блок нужен для разбора спорных кейсов. Здесь видно, какие алиасы использовались, что нашли Ресурс, OCS и Мерлион, и почему поставщик был лучше, дороже или отфильтрован.",
            icon="🧪",
            chips=["debug по источникам", "алиасы артикула", "лучший дистрибьютер и остаток"],
            tone="purple",
        )
        st.markdown("""
        <div class='analysis-help-grid'>
          <div class='analysis-help-card'><div class='analysis-help-title'>🔎 Что смотреть первым</div><div class='analysis-help-text'>Проверь артикул, алиасы и колонку лучшего дистрибьютора. Это быстро покажет, нашёлся ли товар и есть ли цена лучше вашей.</div></div>
          <div class='analysis-help-card'><div class='analysis-help-title'>🧱 Debug по источникам</div><div class='analysis-help-text'>Колонки debug показывают, был ли найден товар у каждого поставщика, дороже он или лучше вас, и какой код реально сработал.</div></div>
          <div class='analysis-help-card'><div class='analysis-help-title'>🛡️ Для чего это</div><div class='analysis-help-text'>Если результат выглядит странно, именно тут проще всего понять: проблема в артикулах, фильтре качества, остатке или просто в цене.</div></div>
        </div>
        """, unsafe_allow_html=True)
        st.dataframe(build_display_df(result_df, price_mode, round100, custom_discount, search_mode=search_mode, min_qty=min_dist_qty), use_container_width=True, hide_index=True, height=360)
    st.download_button(
        "⬇️ Скачать найденное в Excel",
        to_excel_bytes(result_df, price_mode, round100, custom_discount, search_mode=search_mode, min_qty=min_dist_qty),
        file_name="moy_tovar_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    if distributor_sources_ready():
        with st.expander("Показать цены у всех"):
            render_all_distributor_prices_block(result_df, search_mode, min_dist_qty, price_mode, round100, custom_discount)
        render_action_callout(
            "Файл для согласования с руководителем",
            "Этот экспорт собирает базовую аналитику по найденным товарам: ваш текущий прод, лучшую цену поставщика и поля, которые удобно дозаполнить вручную перед обсуждением новых цен.",
            icon="🗂️",
            badges=["артикул и количество уже заполнены", "лучшая цена дистрибьютора уже внутри", "готово для обсуждения"],
        )
        st.download_button(
            "⬇️ Скачать анализ товара",
            build_product_analysis_workbook_bytes(result_df, search_mode, min_qty=min_dist_qty),
            file_name="analysis_product.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Файл для обсуждения новых цен: артикул, количество, текущий прод и лучшая цена дистрибьютора уже заполнены. Остальные поля можно внести вручную.",
        )

st.markdown('</div>', unsafe_allow_html=True)


# ----------------------
# Новый отчёт по прайсу
# ----------------------
st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
render_block_header(
    "Отчёт по всему прайсу",
    "Показывает позиции, где у дистрибьютора цена ниже нашей минимум на выбранный процент. В отчёт попадают только наши позиции в наличии и только нормальные позиции поставщиков в наличии.",
    icon="📦",
    help_text="Массовая аналитика по всему вашему прайсу. Помогает быстро найти товары, по которым поставщики продают заметно дешевле, чем вы сейчас продаёте. Учитывает порог в процентах и минимальный остаток у дистрибьютора.",
)
if current_df is None:
    st.info("Сначала загрузите прайс в левой панели 👈")
elif not distributor_sources_ready():
    st.info("Загрузите хотя бы один файл дистрибьютора: Ресурс, OCS или Мерлион.")
else:
    render_info_banner(
        "Как пользоваться отчётом по прайсу",
        "Этот блок нужен для массового контроля цен. Он показывает только те позиции, где поставщик действительно даёт цену ниже вашей минимум на выбранный процент и при этом проходит фильтр по остатку.",
        icon="📘",
        chips=["поиск по всему прайсу", "учитывает порог выгоды", "отсеивает единичные остатки"],
        tone="blue",
    )
    c1, c2, c3 = st.columns([1, 1, 1.5])
    threshold_val = float(st.session_state.distributor_threshold)
    min_qty_val = float(st.session_state.distributor_min_qty)
    with c1:
        st.metric("Порог", f"{fmt_qty(threshold_val)}%")
    with c2:
        st.metric("Мин. остаток", f"{fmt_qty(min_qty_val)} шт.")
    with c3:
        build_report_clicked = st.button("Показать отчёт", type="primary", use_container_width=True)
    if build_report_clicked:
        st.session_state.distributor_report_df = build_full_distributor_report(current_df, threshold_val, search_mode, min_qty=min_qty_val)

    report_df = st.session_state.get("distributor_report_df")
    if isinstance(report_df, pd.DataFrame) and not report_df.empty:
        render_report_summary_cards(report_df, threshold_val, min_qty_val)
        st.dataframe(report_df, use_container_width=True, hide_index=True, height=440)
        render_action_callout(
            "Экспорт отчёта для работы вне приложения",
            "Скачивай файл, когда нужно быстро обсудить массовые пересмотры цен, отфильтровать позиции в Excel или передать выборку руководителю и коллегам.",
            icon="📥",
            badges=["массовый анализ", "готово для Excel", "цены и остатки уже внутри"],
        )
        st.download_button(
            "⬇️ Скачать отчёт в Excel",
            report_to_excel_bytes(report_df),
            file_name="distributor_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif build_report_clicked:
        st.warning("Ничего не найдено по текущему порогу. Попробуйте снизить % или минимальный остаток.")
st.markdown('</div>', unsafe_allow_html=True)


# ----------
# Авито блок
# ----------
current_avito_df = st.session_state.get("avito_df")
if isinstance(current_avito_df, pd.DataFrame) and not current_avito_df.empty and submitted_query.strip():
    avito_matches = find_avito_ads(current_avito_df, submitted_query, result_df)
    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        "Объявления Авито по этой позиции",
        "Ищу совпадения по введённым артикулам и связанным кодам из найденной позиции. Ссылки читаются прямо из гиперссылок Excel.",
        icon="🛒",
        help_text="Этот блок показывает связанные объявления из файла Авито. Он помогает быстро открыть существующие карточки и проверить, как товар уже продаётся на площадке.",
    )
    if avito_matches.empty:
        st.info("По текущему запросу объявление в файле Авито не найдено.")
    else:
        exact_df = avito_matches[avito_matches["match_kind"] == "exact"].copy()
        related_df = avito_matches[avito_matches["match_kind"] != "exact"].copy()
        if not exact_df.empty:
            st.markdown("**Точные совпадения**")
            for _, row in exact_df.head(30).iterrows():
                with st.container(border=True):
                    c1, c2 = st.columns([5, 1.5])
                    with c1:
                        title = normalize_text(row.get("title", "")) or "Без названия"
                        ad_id = normalize_text(row.get("ad_id", "")) or "Без номера"
                        price = normalize_text(row.get("price", ""))
                        matched = ", ".join(row.get("matched_tokens", []) or [])
                        st.markdown(f"**{title}**")
                        st.caption(f"№ {ad_id}" + (f" • Цена: {price}" if price else ""))
                        if matched:
                            st.caption("Совпали артикулы: " + matched)
                    with c2:
                        render_avito_open_button(str(row.get("url", "")), "Открыть объявление")
        if not related_df.empty:
            st.markdown("**Связанные совпадения**")
            for _, row in related_df.head(30).iterrows():
                with st.container(border=True):
                    c1, c2 = st.columns([5, 1.5])
                    with c1:
                        title = normalize_text(row.get("title", "")) or "Без названия"
                        ad_id = normalize_text(row.get("ad_id", "")) or "Без номера"
                        price = normalize_text(row.get("price", ""))
                        matched = ", ".join(row.get("matched_tokens", []) or [])
                        st.markdown(f"**{title}**")
                        st.caption(f"№ {ad_id}" + (f" • Цена: {price}" if price else ""))
                        if matched:
                            st.caption("Совпали артикулы: " + matched)
                    with c2:
                        render_avito_open_button(str(row.get("url", "")), "Открыть объявление")
    st.markdown('</div>', unsafe_allow_html=True)


# ---------
# Серии
# ---------
series_info = get_series_candidates(current_df, submitted_query, st.session_state.series_mode) if isinstance(current_df, pd.DataFrame) and submitted_query.strip() else {"prefix": "", "candidates": []}
series_candidates = series_info.get("candidates", []) if isinstance(series_info, dict) else []
if current_df is not None and submitted_query.strip() and series_candidates:
    st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
    render_block_header(
        "Серия / цвета по части артикула",
        "Если по части артикула находится серия, можно быстро отметить нужные позиции и добавить их в основной поиск.",
        icon="🎨",
        help_text="Удобный блок для серийных товаров. Если у кода есть несколько цветов, ёмкостей или версий, можно сразу выбрать нужные позиции и одним кликом добавить их в поиск.",
    )
    st.radio("Режим серии", ["Только оригиналы", "Показывать всё"], key="series_mode", horizontal=True)
    series_info = get_series_candidates(current_df, submitted_query, st.session_state.series_mode)
    series_candidates = series_info.get("candidates", []) if isinstance(series_info, dict) else []
    if series_candidates:
        st.caption(f"По префиксу {series_info.get('prefix', '')} найдено позиций: {len(series_candidates)}")
        c_add, c_all, c_clear = st.columns(3)
        prefix_key = normalize_article(str(series_info.get('prefix', '')))
        select_all_clicked = c_all.button("Выбрать все", use_container_width=True, key=f"series_select_all_{prefix_key}")
        clear_all_clicked = c_clear.button("Очистить выбор", use_container_width=True, key=f"series_clear_all_{prefix_key}")
        if select_all_clicked:
            st.session_state[f"series_selected_{prefix_key}"] = [str(c["article_norm"]) for c in series_candidates]
        if clear_all_clicked:
            st.session_state[f"series_selected_{prefix_key}"] = []
        options = [str(c["article_norm"]) for c in series_candidates]
        format_map = {str(c["article_norm"]): f"{c['article']} — свободно: {fmt_qty(c['free_qty'])} • {fmt_price_with_rub(c['sale_price'])} • {c['name']}" for c in series_candidates}
        selected_norms = st.multiselect(
            "Выберите позиции серии",
            options=options,
            default=st.session_state.get(f"series_selected_{prefix_key}", []),
            format_func=lambda x: format_map.get(x, x),
            key=f"series_multiselect_{prefix_key}",
            label_visibility="collapsed",
        )
        st.session_state[f"series_selected_{prefix_key}"] = selected_norms
        add_clicked = c_add.button("Добавить отмеченные в поиск", use_container_width=True, key=f"series_add_{prefix_key}")
        if add_clicked and selected_norms:
            selected_articles = [str(c["article"]) for c in series_candidates if str(c["article_norm"]) in set(selected_norms)]
            normalized_query = "\n".join(unique_preserve_order(selected_articles))
            st.session_state.search_input = normalized_query
            st.session_state.submitted_query = normalized_query
            st.session_state.last_result = perform_search(current_df, normalized_query, search_mode)
            st.rerun()
    else:
        st.info("По этой части артикула серия не найдена или подходящих позиций меньше двух.")
    st.markdown('</div>', unsafe_allow_html=True)


# ----------
# Шаблоны
# ----------
st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
render_block_header(
    "Шаблон 1 — Авито / наличный расчёт",
    "Авито = цена продажи -12%. Наличный = ещё -10% от цены Авито. Если товара нет по «Свободно», будет «продан».",
    icon="🧾",
    help_text="Готовый текст для быстрой отправки или размещения. Автоматически считает цену для Авито и цену за наличный расчёт, а также подставляет статус «продан», если позиции нет в наличии.",
)
if current_df is None:
    st.info("Сначала загрузите прайс в левой панели 👈")
elif not submitted_query.strip():
    st.info("Введите артикулы, затем нажмите **Найти**.")
else:
    template_text = build_offer_template(current_df, submitted_query, round100, st.session_state.template1_footer, search_mode)
    st.session_state["offer_template_text"] = template_text
    line_count = len([x for x in template_text.split("\n\n") if x.strip()]) if template_text.strip() else 0
    st.text_area("Готовый шаблон", height=min(500, max(180, 72 + line_count * 40)), key="offer_template_text")
    if template_text.strip():
        render_copy_big_button(template_text, "📋 Скопировать шаблон 1")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="result-wrap">', unsafe_allow_html=True)
render_block_header(
    "Шаблон 2 — название + выбранная цена",
    f"Цена берётся из выбранного режима слева ({html.escape(price_label)}). Во второй шаблон попадают только позиции, где «Свободно» больше нуля.",
    icon="💵",
    help_text="Короткий шаблон для ценников, сообщений или внутренних согласований. Берёт только позиции в наличии и считает цену по текущему выбранному режиму скидки.",
)
if current_df is None:
    st.info("Сначала загрузите прайс в левой панели 👈")
elif not submitted_query.strip():
    st.info("Введите артикулы, затем нажмите **Найти**.")
else:
    second_template_text = build_selected_price_template(current_df, submitted_query, price_mode, round100, custom_discount, search_mode)
    st.session_state["selected_price_template_text"] = second_template_text
    st.text_area("Готовый шаблон 2", height=min(360, max(150, 52 + max(1, second_template_text.count('\n\n') + 1) * 42)), key="selected_price_template_text")
    if second_template_text.strip():
        render_copy_big_button(second_template_text, "📋 Скопировать шаблон 2")
    else:
        st.info("Во втором шаблоне нечего показывать: найденных позиций в наличии нет.")
st.markdown('</div>', unsafe_allow_html=True)
